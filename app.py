#!/usr/bin/env python3
"""
legistar_sync.py  —  NYC Council legislation tracker (1:1 with Legistar)

Pulls legislation directly from the NYC Legistar Web API (the same data store the
legistar.council.nyc.gov pages render from). Detects what changed since the last run.
Optionally adds AI-drafted impact bullets (cached so re-runs only pay for new/changed bills).

  DATA layer  (exact, from API): Matters, Sponsors, Histories, Attachments, Text.
  ANALYSIS layer (interpretive, labeled): keyword tags and/or AI-drafted impact bullets.

CLI runs ONE profile. For multiple scoped workbooks on a schedule, use run_sync.py + config.json.

  python3 legistar_sync.py --file "Int 0225-2026" --enrich --out int225.xlsx
  python3 legistar_sync.py --since 2024-01-01 --sponsor "<member last name>" --enrich --text --impact ai --out member.xlsx
"""

import argparse, hashlib, json, os, re, sqlite3, sys, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

try:
    import requests
except ImportError:
    requests = None  # allowed; only needed for live runs, not offline tests

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_BASE = "https://webapi.legistar.com/v1/nyc"
# Verified-stable public link: the gateway 302-redirects to the correct LegislationDetail page.
WEB_BASE = "https://legistar.council.nyc.gov/gateway.aspx?m=l&id="
ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
PAGE = 1000


def _ver_rank(v):
    """NYC matter versions are '*' (introduced) then 'A','B',...  Rank '*' lowest, letters ascending.
    (The stock scraper assumes integer versions and would crash on NYC's letters.)"""
    if v is None:
        return (-1, "")
    return (0, "") if str(v) == "*" else (1, str(v))


# ============================================================================
# DATA LAYER — Legistar Web API client
# ============================================================================
class LegistarClient:
    def __init__(self, token=None, pause=0.0):
        self.token = token or os.environ.get("LEGISTAR_TOKEN")
        self.pause = pause
        self.s = requests.Session()
        self.s.headers.update({"Accept": "application/json"})

    def _get(self, path, params=None):
        params = dict(params or {})
        if self.token:
            params["token"] = self.token
        last_err = None
        for attempt in range(6):
            try:
                r = self.s.get(f"{API_BASE}/{path}", params=params, timeout=90)
            except requests.exceptions.RequestException as e:
                last_err = e
                time.sleep(min(2 ** attempt, 20))  # ride out dropped connections
                continue
            if r.status_code == 200:
                if self.pause:
                    time.sleep(self.pause)
                return r.json()
            if r.status_code in (429, 500, 502, 503, 504):
                time.sleep(min(2 ** attempt, 20)); continue
            r.raise_for_status()
        if last_err:
            raise last_err
        r.raise_for_status()

    def matters(self, odata_filter=None, top=PAGE):
        # Page WITHOUT a volatile $orderby and de-dupe by MatterId, exactly as the
        # production NYC scraper does: ordering by a non-unique field while paging by
        # $skip can drop/duplicate rows at page boundaries.
        out, seen, skip = [], set(), 0
        while True:
            params = {"$top": top, "$skip": skip}
            if odata_filter:
                params["$filter"] = odata_filter
            batch = self._get("matters", params)
            for m in batch:
                mid = m.get("MatterId")
                if mid not in seen:
                    seen.add(mid)
                    out.append(m)
            if len(batch) < top:
                return out
            skip += top

    def sponsors(self, mid):     return self._get(f"matters/{mid}/sponsors")
    def histories(self, mid):    return self._get(f"matters/{mid}/histories")
    def attachments(self, mid):  return self._get(f"matters/{mid}/attachments")

    def text_plain(self, mid, matter_version=None):
        # Verified shape: /matters/{id}/versions -> [{"Key":<text id>, "Value":<'*'|'A'|...>}],
        # pick the version whose Value == the matter's MatterVersion (else highest), then
        # /matters/{id}/texts/{Key} -> MatterTextPlain.
        try:
            versions = self._get(f"matters/{mid}/versions")
        except Exception:
            return ""
        if not versions:
            return ""
        chosen = None
        if matter_version is not None:
            chosen = next((v for v in versions if str(v.get("Value")) == str(matter_version)), None)
        if chosen is None:
            chosen = max(versions, key=lambda v: _ver_rank(v.get("Value")))
        try:
            t = self._get(f"matters/{mid}/texts/{chosen.get('Key')}")
            if isinstance(t, list):
                t = t[-1] if t else {}
            return (t or {}).get("MatterTextPlain") or ""
        except Exception:
            return ""

    def events(self, odata_filter=None, top=PAGE):
        out, seen, skip = [], set(), 0
        while True:
            params = {"$top": top, "$skip": skip}
            if odata_filter:
                params["$filter"] = odata_filter
            batch = self._get("events", params)
            for e in batch:
                eid = e.get("EventId")
                if eid not in seen:
                    seen.add(eid)
                    out.append(e)
            if len(batch) < top:
                return out
            skip += top

    def event_items(self, event_id):
        try:
            return self._get(f"events/{event_id}/eventitems")
        except Exception:
            return []

    def votes(self, history_id):
        # NYC quirk: roll-call votes are fetched via the matter-history id
        try:
            return self._get(f"eventitems/{history_id}/votes")
        except Exception:
            return []

    def council_members(self):
        """Current Council Member names from the City Council body's active office records."""
        try:
            bodies = self._get("bodies")
            cc = next((b for b in bodies if (b.get("BodyName") or "").strip().lower() == "city council"), None)
            if cc:
                recs = self._get(f"bodies/{cc['BodyId']}/OfficeRecords")
                today = datetime.now().date().isoformat()
                names = set()
                for r in recs:
                    end = (r.get("OfficeRecordEndDate") or "")[:10]
                    if not end or end >= today:
                        nm = (r.get("OfficeRecordPersonName") or r.get("OfficeRecordFullName") or "").strip()
                        if nm:
                            names.add(nm)
                if names:
                    return sorted(names)
        except Exception:
            pass
        try:
            persons = self._get("persons", {"$filter": "PersonActiveFlag eq 1"})
            return sorted({(p.get("PersonFullName") or "").strip() for p in persons if p.get("PersonFullName")})
        except Exception:
            return []

    def committees(self):
        """Current committees with their chair + members, from bodies/OfficeRecords.

        These are the official gatekeepers for moving a bill — used to ground the
        Influence Map in who actually chairs and sits on each committee.
        """
        try:
            bodies = self._get("bodies")
        except Exception:
            return []
        coms = [b for b in bodies if "committee" in (b.get("BodyTypeName") or "").lower()
                or "committee" in (b.get("BodyName") or "").lower()]
        today = datetime.now().date().isoformat()

        def _one(b):
            try:
                recs = self._get(f"bodies/{b['BodyId']}/OfficeRecords")
            except Exception:
                return None
            chair, members = "", []
            for r in recs or []:
                end = (r.get("OfficeRecordEndDate") or "")[:10]
                if end and end < today:
                    continue
                nm = (r.get("OfficeRecordPersonName") or r.get("OfficeRecordFullName") or "").strip()
                if not nm:
                    continue
                title = (r.get("OfficeRecordTitle") or "").lower()
                if "chair" in title and "vice" not in title and not chair:
                    chair = nm
                members.append(nm)
            if not members:
                return None
            return {"committee": (b.get("BodyName") or "").strip(), "chair": chair,
                    "members": sorted(set(members))}

        out = []
        with ThreadPoolExecutor(max_workers=8) as ex:
            for fut in as_completed([ex.submit(_one, b) for b in coms]):
                try:
                    r = fut.result()
                    if r:
                        out.append(r)
                except Exception:
                    pass
        out.sort(key=lambda x: x["committee"])
        return out


# ============================================================================
# DATA LAYER — pure transforms (no network; unit-testable)
# ============================================================================
def web_url(m):
    return f"{WEB_BASE}{m.get('MatterId')}"


def current_sponsors(matter, sponsor_json):
    # Mirror the production scraper: keep sponsors at the latest version, sorted by sequence.
    # Prefer the matter's MatterVersion; else the highest sponsor version (NYC-safe rank).
    rows = [s for s in (sponsor_json or []) if (s.get("MatterSponsorName") or "").strip()]
    if not rows:
        return []
    target = matter.get("MatterVersion")
    present = {s.get("MatterSponsorMatterVersion") for s in rows}
    if target not in present:
        target = max(present, key=_ver_rank)
    scoped = [s for s in rows if s.get("MatterSponsorMatterVersion") == target] or rows
    by_name = {}
    for s in scoped:
        name = s["MatterSponsorName"].strip()
        seq = s.get("MatterSponsorSequence", 9999)
        if name not in by_name or seq < by_name[name].get("MatterSponsorSequence", 9999):
            by_name[name] = s
    return sorted(by_name.values(), key=lambda s: s.get("MatterSponsorSequence", 9999))


def _date(s):
    if not s:
        return ""
    try:
        return datetime.fromisoformat(str(s).replace("Z", "")).strftime("%-m/%-d/%Y")
    except Exception:
        return str(s)[:10]


def normalize_event(e):
    insite = e.get("EventInSiteURL") or f"https://legistar.council.nyc.gov/gateway.aspx?m=e&id={e.get('EventId')}"
    return {
        "Date": _date(e.get("EventDate")),
        "Time": (e.get("EventTime") or "").strip(),
        "Committee / Body": e.get("EventBodyName", ""),
        "Location": (e.get("EventLocation") or "").strip(),
        "Agenda status": e.get("EventAgendaStatusName", ""),
        "Agenda": e.get("EventAgendaFile", "") or "",
        "Minutes": e.get("EventMinutesFile", "") or "",
        "Legistar": insite,
        "EventId": e.get("EventId"),
        "_sortdate": e.get("EventDate") or "",
    }


# --- NYC Open Data (311) grounding for bill analysis ---
COMPLAINT_MAP = [
    (("noise", "amplified", "quiet"), "Noise"),
    (("rat", "rodent", "vermin", "pest"), "Rodent"),
    (("tree", "forestry"), "Damaged Tree"),
    (("pothole", "roadway", "pavement", "street condition"), "Street Condition"),
    (("heat", "hot water", "boiler"), "HEAT/HOT WATER"),
    (("homeless",), "Homeless Person Assistance"),
    (("tow", "abandoned vehicle", "illegal park", "parking"), "Illegal Parking"),
    (("garbage", "trash", "litter", "dumping", "sanitation"), "Dirty Condition"),
    (("flood", "sewer", "storm", "drain", "catch basin"), "Sewer"),
    (("graffiti",), "Graffiti"),
    (("idling", "emissions", "air quality"), "Air Quality"),
    (("sidewalk", "scaffold", "construction"), "Sidewalk Condition"),
    (("lead", "mold", "asbestos"), "Lead"),
]
_BOROUGHS = ["Manhattan", "Brooklyn", "Queens", "Bronx", "Staten Island"]


def matter_vote_events(client, histories):
    """Collect roll-call vote events for a matter from its histories.
    Returns a list of {date, action, body, result, tally{value:count}, votes[{Member,Vote}]}."""
    out = []
    for h in histories or []:
        if not h.get("MatterHistoryRollCallFlag"):
            continue
        hid = h.get("MatterHistoryId")
        if hid is None:
            continue
        vs = client.votes(hid)
        if not vs:
            continue
        tally, rows = {}, []
        for v in vs:
            val = (v.get("VoteValueName") or "").strip() or "—"
            nm = (v.get("VotePersonName") or "").strip()
            tally[val] = tally.get(val, 0) + 1
            rows.append({"Member": nm, "Vote": val})
        out.append({
            "date": _date(h.get("MatterHistoryActionDate")),
            "action": (h.get("MatterHistoryActionName") or "").strip(),
            "body": (h.get("MatterHistoryActionBodyName") or "").strip(),
            "result": (h.get("MatterHistoryPassedFlagName") or "").strip(),
            "tally": tally,
            "votes": sorted(rows, key=lambda r: (r["Vote"], r["Member"])),
        })
    return out


def bill_311_keyword(row):
    blob = ((row.get("Title") or "") + " " + (row.get("Name") or "")).lower()
    for terms, ctype in COMPLAINT_MAP:
        if any(t in blob for t in terms):
            return ctype
    return None


def nyc_311_context(borough=None, complaint_type=None, days=365, timeout=20):
    """Live count of 311 complaints from NYC Open Data (dataset erm2-nwe9). Best-effort."""
    from datetime import timedelta
    import urllib.request
    import urllib.parse
    base = "https://data.cityofnewyork.us/resource/erm2-nwe9.json"
    since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%dT00:00:00")
    where = [f"created_date >= '{since}'"]
    if borough:
        where.append(f"upper(borough)='{borough.upper()}'")
    if complaint_type:
        where.append(f"complaint_type='{complaint_type}'")
    url = base + "?" + urllib.parse.urlencode({"$select": "count(*) AS n", "$where": " AND ".join(where)})
    try:
        with urllib.request.urlopen(url, timeout=timeout) as resp:
            data = json.loads(resp.read().decode())
        return {"ok": True, "count": (data[0].get("n") if data else None), "url": url}
    except Exception as e:
        return {"ok": False, "error": str(e), "url": url}


def build_data_context(row):
    """Assemble a small REAL-DATA block for a bill (currently 311). Returns '' if nothing relevant."""
    ctype = bill_311_keyword(row)
    if not ctype:
        return ""
    bors = [b for b in _BOROUGHS if b in (row.get("Boroughs named") or "")]
    bor = bors[0] if len(bors) == 1 else None
    scope = bor or "citywide"
    res = nyc_311_context(bor, ctype)
    if res.get("ok") and res.get("count") is not None:
        return (f"NYC 311 service requests, last 12 months, {scope}: {res['count']} '{ctype}' complaints. "
                f"Source: NYC Open Data 311 (dataset erm2-nwe9).")
    return (f"A NYC 311 lookup for '{ctype}' complaints ({scope}) is relevant here but was not retrieved live; "
            f"verify at NYC Open Data (dataset erm2-nwe9).")


def normalize_matter(m, sponsors=None, histories=None, attachments=None):
    sp = current_sponsors(m, sponsors) if sponsors is not None else None
    prime = next((s["MatterSponsorName"] for s in (sp or []) if s.get("MatterSponsorSequence") == 0), "")
    if not prime and sp:
        prime = sp[0]["MatterSponsorName"]
    names = [s["MatterSponsorName"] for s in (sp or [])]
    latest = None
    if histories:
        dated = [h for h in histories if h.get("MatterHistoryActionDate")]
        latest = max(dated, key=lambda h: h["MatterHistoryActionDate"], default=None)
    return {
        "File": m.get("MatterFile") or "", "Type": m.get("MatterTypeName") or "",
        "Name": m.get("MatterName") or "", "Title": m.get("MatterTitle") or "",
        "Status": m.get("MatterStatusName") or "", "Committee/Body": m.get("MatterBodyName") or "",
        "Intro Date": _date(m.get("MatterIntroDate")), "Agenda Date": _date(m.get("MatterAgendaDate")),
        "Passed Date": _date(m.get("MatterPassedDate")), "Enacted Date": _date(m.get("MatterEnactmentDate")),
        "Law #": m.get("MatterEnactmentNumber", ""),
        "Sponsors (#)": len(names) if sp is not None else "",
        "Prime Sponsor": prime,
        "Latest Action": (latest or {}).get("MatterHistoryActionName", ""),
        "Latest Action Date": _date((latest or {}).get("MatterHistoryActionDate")),
        "Attachments (#)": len(attachments) if attachments is not None else "",
        "Last Modified (UTC)": m.get("MatterLastModifiedUtc") or "",
        "Version": m.get("MatterVersion") or "*",
        "MatterId": m.get("MatterId", ""), "Web Link": web_url(m),
        "_sponsor_names": names, "_prime": prime,
        "_sponsor_objs": sp or [], "_status_raw": m.get("MatterStatusName", ""),
        "_intro_raw": m.get("MatterIntroDate") or "",
    }


# ============================================================================
# ANALYSIS LAYER (1) — heuristic policy-topic tagging, citywide (free, transparent)
# ============================================================================
TOPICS = {
    "Housing & Buildings": ["housing", "affordable housing", "tenant", "landlord", "rent", "rental", "rent stabilization",
        "HPD", "construction", "NYCHA", "public housing", "homeownership", "mortgage", "foreclosure", "lead paint",
        "eviction", "SRO", "basement apartment", "certificate of occupancy", "building code"],
    "Land Use & Zoning": ["zoning", "rezoning", "rezone", "land use", "ULURP", "landmark", "preservation",
        "city planning", "special district", "waterfront", "variance", "comprehensive plan", "FAR"],
    "Transportation & Streets": ["transportation", "DOT", "bus", "buses", "bus lane", "subway", "MTA", "bike", "bicycle",
        "e-bike", "scooter", "pedestrian", "traffic", "parking", "ferry", "speed limit", "vision zero",
        "congestion pricing", "sidewalk", "crosswalk", "taxi", "for-hire vehicle", "bus stop"],
    "Public Safety & Policing": ["NYPD", "police", "policing", "crime", "public safety", "gun", "guns", "firearm",
        "surveillance", "body camera", "precinct", "hate crime", "domestic violence", "shooting", "ghost gun"],
    "Fire & Emergency Management": ["FDNY", "fire", "EMS", "emergency", "911", "evacuation", "disaster",
        "emergency management", "first responder"],
    "Criminal Justice & Courts": ["jail", "Rikers", "incarceration", "correction", "DOC", "reentry", "bail",
        "district attorney", "court", "justice system", "parole", "probation", "solitary", "detainee"],
    "Health & Mental Health": ["health", "hospital", "H+H", "clinic", "medical", "DOHMH", "mental health", "substance use",
        "overdose", "opioid", "disease", "vaccine", "maternal", "reproductive", "public health", "naloxone",
        "lead poisoning"],
    "Social Services & Homelessness": ["homeless", "shelter", "DHS", "HRA", "public assistance", "SNAP", "benefits",
        "food insecurity", "hunger", "poverty", "cash assistance", "general welfare", "social services", "food pantry"],
    "Education & Schools": ["education", "school", "DOE", "student", "teacher", "classroom", "curriculum", "CUNY",
        "college", "pre-k", "3-K", "special education", "charter school", "literacy", "tutoring"],
    "Children, Youth & Families": ["youth", "children", "ACS", "child welfare", "foster care", "daycare", "childcare",
        "after-school", "summer youth", "family", "juvenile", "early childhood"],
    "Aging & Older Adults": ["aging", "senior", "older adult", "DFTA", "elderly", "NORC", "aging in place", "senior center"],
    "Immigration": ["immigrant", "immigration", "MOIA", "asylum", "refugee", "undocumented", "IDNYC", "language access",
        "naturalization", "newcomer"],
    "Economic & Small Business": ["economic development", "EDC", "small business", "commercial", "workforce",
        "employment", "jobs", "entrepreneur", "MWBE", "minority-owned", "women-owned", "industrial", "manufacturing",
        "storefront", "tourism"],
    "Labor & Workers": ["labor", "worker", "wage", "minimum wage", "paid leave", "sick leave", "union",
        "collective bargaining", "gig worker", "freelance", "prevailing wage", "workplace safety", "delivery worker"],
    "Consumer & Worker Protection": ["consumer", "DCWP", "license", "debt", "predatory", "deceptive", "price gouging",
        "tenant harassment", "worker protection", "licensing"],
    "Environment & Sanitation": ["environment", "climate", "emissions", "sustainability", "resiliency", "flood", "storm",
        "sewer", "water", "air quality", "DEP", "sanitation", "DSNY", "waste", "recycling", "composting", "rats",
        "rodent", "litter", "local law 97", "green infrastructure", "stormwater", "solar"],
    "Parks & Open Space": ["parks", "recreation", "playground", "open space", "DPR", "trees", "tree planting",
        "urban forest", "garden", "greenway", "waterfront access", "park land"],
    "Arts, Culture & Libraries": ["cultural", "museum", "artist", "theater", "library", "libraries", "heritage",
        "public art", "DCLA", "humanities", "historic", "gallery", "arts education"],
    "Technology & Privacy": ["technology", "broadband", "internet", "data privacy", "privacy", "cybersecurity", "OTI",
        "algorithm", "artificial intelligence", "digital", "open data", "smart city"],
    "Civil Rights & Equity": ["civil rights", "discrimination", "human rights", "CCHR", "LGBTQ", "disability",
        "accessibility", "gender", "racial equity", "bias", "equity", "reproductive rights"],
    "Veterans": ["veteran", "veterans", "DVS", "military", "armed forces", "servicemember"],
    "Government & Elections": ["election", "elections", "voting", "voter", "ballot", "campaign finance", "ethics", "COIB",
        "transparency", "FOIL", "open meeting", "charter", "redistricting", "governmental operations", "lobbying",
        "board of elections", "ranked choice"],
    "Budget, Finance & Taxes": ["budget", "fiscal", "tax", "taxes", "revenue", "appropriation", "OMB", "comptroller",
        "audit", "property tax", "tax exemption", "tax abatement", "capital budget"],
    "Contracts & Procurement": ["contract", "procurement", "MOCS", "vendor", "RFP", "prompt payment",
        "nonprofit contract", "subcontract"],
}
PILLARS = TOPICS  # backward-compat alias
BOROUGHS = ["Manhattan", "Brooklyn", "Queens", "Bronx", "Staten Island"]


def _kw_hit(blob, k):
    k = k.lower()
    if re.fullmatch(r"[a-z0-9 ]+", k):  # plain words/phrases → word-boundary match, optional plural
        return re.search(r"\b" + re.escape(k) + r"(?:s|es)?\b", blob) is not None
    return k in blob  # has punctuation (H+H, 3-K, pre-k, e-bike) → substring


def keyword_tags(row, text=""):
    blob = " ".join([row.get("Title", ""), row.get("Name", ""), text]).lower()
    topics = [t for t, kws in TOPICS.items() if any(_kw_hit(blob, k) for k in kws)]
    bor = [b for b in BOROUGHS if b.lower() in blob]
    return {"Topic Tags": "; ".join(topics),
            "Boroughs Named": "; ".join(bor) or "(citywide / none named)"}


# ============================================================================
# ANALYSIS LAYER (2) — AI-drafted impact bullets (Anthropic API, cached)
# ============================================================================
PROMPT = """You are a nonpartisan legislative analyst for the New York City Council. Given a NYC Council \
bill's file number, title, and text, write a tight, citywide impact read. Ground every statement ONLY in the \
provided text; do not invent specifics. Treat every district and Council Member equally.

Return ONLY a JSON object with keys "purpose", "affects", "local":
- purpose: <=25 words, plain language, what the bill actually does.
- affects: <=30 words, who/what across the city it touches and how.
- local: <=35 words, where the bill's effects concentrate geographically (which boroughs/neighborhoods/districts, \
if any), or "citywide" if it applies uniformly. Be specific only where the text supports it; do not stretch.

File: {file}
Type: {type}
Title: {title}
Text: {text}

JSON:"""


DOSSIER_PROMPT = """You are a nonpartisan legislative analyst preparing a neutral staff profile of a sitting \
NYC Council Member, for internal professional use. Use ONLY the structured data provided below (their sponsorship \
record, topic mix, bill outcomes, and frequent co-sponsors).

STRICT RULES:
- Do NOT invent or assume facts: no quotes, biographical details, party label, district, scandals, endorsements, \
or policy positions that are not derivable from the data.
- If the data is thin or a section can't be supported, say so plainly rather than speculating.
- Neutral, factual, analytical tone. No praise, no criticism, no campaign language.
- Clearly mark interpretive statements as inferences ("the record suggests...").

Write ~180 words with these short labeled sections:
1. Policy focus — the leading topic areas, from the topic mix.
2. Legislative activity — volume, prime vs co-sponsor balance, and how much has passed vs stalled.
3. Coalition — the colleagues they most often co-sponsor alongside.
4. Patterns & caveats — any notable pattern, plus a one-line caveat that this reflects sponsorship activity only \
(not floor votes) and is AI-generated.

Member: {member}
Legislative data (JSON): {stats}

Profile:"""


ANALYSIS_PROMPT = """You are a nonpartisan legislative policy analyst writing an internal briefing on a NYC Council bill.
Use the bill information and any REAL DATA CONTEXT provided below.

RULES:
- Ground every claim in the bill text or the provided data. Do NOT invent statistics, dollar amounts, poll numbers,
  agency budgets, or quotes. If you lack a figure, say what data would answer it and name the source to check.
- Be balanced: present support and opposition fairly. Mark predictions clearly as predictions.
- Neutral, professional tone. Label the brief as AI analysis/inference, not an official statement.

Write these sections, each a bold header followed by 2–4 substantive bullets (be specific, not generic):
**What the bill does** — the mechanism in plain language, what legally changes, who it covers, and effective date/triggers if stated.
**Who would support it** — specific constituencies, agencies, advocacy groups, and the concrete reasons they'd back it.
**Who would oppose it / concerns** — likely opponents, their strongest objections, trade-offs, and implementation friction.
**Political analysis** — sponsor coalition, partisan and borough dynamics, committee path, and what moving it realistically takes (votes, Speaker, mayoral posture).
**District / borough / citywide outlook** — how effects differ at the district, borough, and citywide levels, and which areas are most affected.
**Fiscal analysis** — qualitative cost and revenue drivers, who bears the cost, enforcement burden, and what an OMB/IBO fiscal note would examine; name the
  specific figures to verify. Do NOT fabricate numbers.
**Why it exists / who needed it** — the underlying problem and its scale; cite the REAL DATA CONTEXT if present, and name the NYC
  Open Data / agency datasets (311, HPD, DOB, DOT, NYPD, DOHMH, IBO) that would document the need.
**Comparable measures** — similar laws/proposals in NYC's past or other cities, and how they fared (reason from general knowledge; mark as illustrative).
**If implemented** — near-term and longer-term effects, second-order consequences, risks, and concrete metrics to monitor afterward.
**Open questions** — the 2–3 sharpest questions a staffer should resolve before the sponsor commits.

BILL
File: {file} | Type: {type} | Status: {status} | Committee: {committee}
Title: {title}
Sponsors: {sponsors}
Text (excerpt): {text}

REAL DATA CONTEXT (may be empty):
{data}

Briefing:"""


QUERY_PLAN_PROMPT = """Convert the user's question about NYC City Council bills into a JSON execution plan.
Return ONLY a JSON object (no prose, no markdown fences) with these keys:
- "intent": one of "filter" (list/find bills), "compare" (compare named bills), "explain" (explain one bill), "answer" (general question about the loaded set).
- "bill_refs": array of bill identifiers the user names, e.g. ["220","Int 0419-2026"]. Empty if none.
- "filters": object; include only keys that apply: "keyword" (free text), "topic" (a policy area keyword such as housing, land use, transportation, public safety, fire, criminal justice, health, social services, education, youth, aging, immigration, economic development, labor, consumer protection, environment, sanitation, parks, arts, technology, civil rights, veterans, elections, budget, or contracts), "type" (e.g. Introduction, Resolution, Land Use Application), "status" (e.g. Committee, Enacted), "borough" (Manhattan/Brooklyn/Queens/Bronx/Staten Island), "sponsor" (a member name), "min_sponsors" (integer).
- "needs_detail": true if answering needs the full text/sponsors of the named bills (comparisons and explanations do).

User question: {q}
JSON:"""


QUERY_ANSWER_PROMPT = """You are an expert assistant on New York City government, law, and legislation, helping a \
NYC City Council staffer. Answer like a knowledgeable, plain-spoken legislative analyst. You can handle:
- Questions about NYC Council bills — use the EVIDENCE provided (cite file numbers like Int 0220-2026).
- How city government works — the City Charter, the Administrative Code, the Rules of the City of New York (RCNY), the
  legislative process, committees, ULURP / land use, the budget and capital process, oversight, and what agencies do.
- Finding the actual law — when asked for a code section or ordinance, give the specific citation (e.g.,
  "Admin. Code § 27-2005", "Charter § 197-c", "1 RCNY") and, when web search is available, look it up and link an
  authoritative source (nyc.gov, the NYC Administrative Code / American Legal Publishing, the NY State Senate/Legislature).

Rules:
- For facts about specific loaded bills, use ONLY the EVIDENCE; never invent sponsors, counts, status, or text.
- For law and process questions, be accurate and cite the specific section. If you are not certain of a citation,
  search for it or say you're unsure — NEVER fabricate a section number or quote.
- Prefer official/primary sources. Be concise and practical; use short paragraphs and bullets.

USER QUESTION: {q}

EVIDENCE from the loaded NYC Council data (may be empty or not relevant for general questions):
{ev}

Answer:"""


def text_hash(*parts):
    return hashlib.sha1("||".join(p or "" for p in parts).encode("utf-8", "ignore")).hexdigest()[:16]


class AIImpact:
    def __init__(self, model, api_key=None, pause=0.0):
        self.model = model
        self.key = api_key or os.environ.get("ANTHROPIC_API_KEY")
        self.pause = pause
        self.s = requests.Session() if requests else None

    def draft(self, row, text):
        if not self.key:
            raise RuntimeError("ANTHROPIC_API_KEY not set")
        prompt = PROMPT.format(file=row.get("File", ""), type=row.get("Type", ""),
                               title=row.get("Title", ""), text=((text or row.get("Name") or ""))[:6000])
        body = {"model": self.model, "max_tokens": 320,
                "messages": [{"role": "user", "content": prompt}]}
        r = self.s.post(ANTHROPIC_URL, headers={
            "x-api-key": self.key, "anthropic-version": "2023-06-01",
            "content-type": "application/json"}, json=body, timeout=90)
        r.raise_for_status()
        data = r.json()
        txt = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
        if self.pause:
            time.sleep(self.pause)
        return parse_bullets(txt)

    def dossier(self, member, stats):
        if not self.key:
            raise RuntimeError("ANTHROPIC_API_KEY not set")
        prompt = DOSSIER_PROMPT.format(member=member, stats=json.dumps(stats, ensure_ascii=False)[:6000])
        body = {"model": self.model, "max_tokens": 700,
                "messages": [{"role": "user", "content": prompt}]}
        r = self.s.post(ANTHROPIC_URL, headers={
            "x-api-key": self.key, "anthropic-version": "2023-06-01",
            "content-type": "application/json"}, json=body, timeout=120)
        r.raise_for_status()
        data = r.json()
        return "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text").strip()

    def analyze(self, row, text, data_ctx=""):
        if not self.key:
            raise RuntimeError("ANTHROPIC_API_KEY not set")
        sponsors = ", ".join(s.get("MatterSponsorName", "") for s in (row.get("_sponsor_objs") or [])) or "(not loaded)"
        prompt = ANALYSIS_PROMPT.format(
            file=row.get("File", ""), type=row.get("Type", ""), status=row.get("Status", ""),
            committee=row.get("Committee/Body", ""), title=row.get("Title", ""),
            sponsors=sponsors, text=((text or row.get("Name") or ""))[:7000], data=data_ctx or "(none retrieved)")
        body = {"model": self.model, "max_tokens": 1900,
                "messages": [{"role": "user", "content": prompt}]}
        r = self.s.post(ANTHROPIC_URL, headers={
            "x-api-key": self.key, "anthropic-version": "2023-06-01",
            "content-type": "application/json"}, json=body, timeout=150)
        r.raise_for_status()
        data = r.json()
        return "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text").strip()

    def chat_plan(self, q):
        if not self.key:
            raise RuntimeError("ANTHROPIC_API_KEY not set")
        body = {"model": self.model, "max_tokens": 400,
                "messages": [{"role": "user", "content": QUERY_PLAN_PROMPT.format(q=q[:1500])}]}
        r = self.s.post(ANTHROPIC_URL, headers={
            "x-api-key": self.key, "anthropic-version": "2023-06-01",
            "content-type": "application/json"}, json=body, timeout=60)
        r.raise_for_status()
        txt = "".join(b.get("text", "") for b in r.json().get("content", []) if b.get("type") == "text")
        try:
            return json.loads(txt)
        except Exception:
            m = re.search(r"\{.*\}", txt, re.DOTALL)
            try:
                return json.loads(m.group(0)) if m else {}
            except Exception:
                return {}

    def chat_answer(self, q, evidence, allow_web=True):
        if not self.key:
            raise RuntimeError("ANTHROPIC_API_KEY not set")
        ev = json.dumps(evidence, ensure_ascii=False)[:8000]
        body = {"model": self.model, "max_tokens": 1600,
                "messages": [{"role": "user", "content": QUERY_ANSWER_PROMPT.format(q=q[:2000], ev=ev)}]}
        if allow_web:
            body["tools"] = [{"type": "web_search_20250305", "name": "web_search"}]
        r = self.s.post(ANTHROPIC_URL, headers={
            "x-api-key": self.key, "anthropic-version": "2023-06-01",
            "content-type": "application/json"}, json=body, timeout=180)
        r.raise_for_status()
        return "".join(b.get("text", "") for b in r.json().get("content", []) if b.get("type") == "text").strip()


def parse_bullets(txt):
    s = re.sub(r"^```(json)?|```$", "", (txt or "").strip(), flags=re.MULTILINE).strip()
    try:
        d = json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        d = json.loads(m.group(0)) if m else {}
    return {"purpose": d.get("purpose", ""), "affects": d.get("affects", ""), "local": d.get("local", d.get("d49", ""))}


# ============================================================================
# CHANGE DETECTION + caches — local snapshot
# ============================================================================
def fingerprint(row):
    return {"File": row["File"], "Status": row["Status"], "SponsorCount": row["Sponsors (#)"],
            "Sponsors": sorted(row.get("_sponsor_names", [])), "Prime": row.get("_prime", ""),
            "Attachments": row["Attachments (#)"], "LatestAction": row["Latest Action"],
            "LastModified": row["Last Modified (UTC)"], "Version": row.get("Version", "")}


class Snapshot:
    def __init__(self, path="legistar_state.db"):
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        self.db = sqlite3.connect(path)
        self.db.execute("CREATE TABLE IF NOT EXISTS state (matter_id INTEGER PRIMARY KEY, fp TEXT)")
        self.db.execute("CREATE TABLE IF NOT EXISTS ai (matter_id INTEGER, thash TEXT, model TEXT, "
                        "bullets TEXT, created TEXT, PRIMARY KEY (matter_id, thash, model))")
        self.db.commit()

    def load(self):
        return {mid: json.loads(fp) for mid, fp in self.db.execute("SELECT matter_id, fp FROM state")}

    def save(self, rows):
        self.db.executemany("INSERT OR REPLACE INTO state VALUES (?,?)",
                            [(r["MatterId"], json.dumps(fingerprint(r))) for r in rows])
        self.db.commit()

    def ai_get(self, mid, thash, model):
        row = self.db.execute("SELECT bullets FROM ai WHERE matter_id=? AND thash=? AND model=?",
                              (mid, thash, model)).fetchone()
        return json.loads(row[0]) if row else None

    def ai_put(self, mid, thash, model, bullets):
        self.db.execute("INSERT OR REPLACE INTO ai VALUES (?,?,?,?,?)",
                        (mid, thash, model, json.dumps(bullets),
                         datetime.now(timezone.utc).isoformat()))
        self.db.commit()


def diff(old, rows):
    changes = []
    for r in rows:
        mid, new = r["MatterId"], fingerprint(r)
        when = (r.get("Last Modified (UTC)") or "")[:16].replace("T", " ")
        prev = old.get(mid)
        if prev is None:
            changes.append((r["File"], "NEW", "—", "—", f"{r['Type']}: {(r['Title'] or '')[:80]}", when)); continue
        if prev["Status"] != new["Status"]:
            changes.append((r["File"], "STATUS", "Status", prev["Status"], new["Status"], when))
        if prev.get("Version") != new.get("Version"):
            changes.append((r["File"], "AMENDED (text)", "Version", prev.get("Version", ""), new.get("Version", ""), when))
        added = set(new["Sponsors"]) - set(prev["Sponsors"])
        dropped = set(prev["Sponsors"]) - set(new["Sponsors"])
        if added:
            changes.append((r["File"], "SIGNED ON +", "Sponsors", f'{prev["SponsorCount"]} sponsors', "; ".join(sorted(added)), when))
        if dropped:
            changes.append((r["File"], "REMOVED -", "Sponsors", "; ".join(sorted(dropped)), f'now {new["SponsorCount"]}', when))
        if prev["Prime"] != new["Prime"]:
            changes.append((r["File"], "PRIME", "Prime Sponsor", prev["Prime"], new["Prime"], when))
        if prev["LatestAction"] != new["LatestAction"]:
            changes.append((r["File"], "ACTION", "Latest Action", prev["LatestAction"], new["LatestAction"], when))
        if prev["Attachments"] != new["Attachments"]:
            changes.append((r["File"], "ATTACH", "Attachments (#)", prev["Attachments"], new["Attachments"], when))
    return changes


# ============================================================================
# OUTPUT — house-style workbook
# ============================================================================
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)
AI_FONT = Font(name="Arial", size=10, italic=True)
CHANGE_FILL = PatternFill("solid", fgColor="FFF2CC")
AI_FILL = PatternFill("solid", fgColor="EAF1FB")
THIN = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _sheet(wb, title, headers, rows, widths=None, fill=None, font=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c); cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = font or BODY; cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for i, w in enumerate(widths or [], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def build_workbook(bundle, path):
    rows = bundle["rows"]
    sponsors_map, histories_map = bundle["sponsors_map"], bundle["histories_map"]
    attach_map, text_map = bundle["attach_map"], bundle["text_map"]
    ai_map, changes, run_info = bundle["ai_map"], bundle["changes"], bundle["run_info"]
    impact_mode = bundle["impact_mode"]

    wb = Workbook(); wb.remove(wb.active)

    cols = ["File", "Type", "Title", "Status", "Committee/Body", "Intro Date", "Latest Action",
            "Latest Action Date", "Sponsors (#)", "Prime Sponsor", "Law #",
            "Last Modified (UTC)", "MatterId", "Web Link"]
    add_kw = impact_mode in ("keyword", "ai")
    if add_kw:
        cols += ["Topic Tags", "Boroughs Named"]
    body = []
    for r in rows:
        line = [r.get(c, "") for c in cols if c not in ("Topic Tags", "Boroughs Named")]
        if add_kw:
            t = keyword_tags(r, (text_map.get(r["MatterId"], "") or "")[:4000])
            line += [t["Topic Tags"], t["Boroughs Named"]]
        body.append(line)
    _sheet(wb, "Matters", cols, body,
           widths=[16, 15, 58, 15, 24, 11, 26, 12, 11, 22, 9, 22, 10, 50] + ([24, 22] if add_kw else []))

    if impact_mode == "ai":
        ai_rows = []
        for r in rows:
            b = ai_map.get(r["MatterId"])
            if b:
                ai_rows.append([r["File"], r["Title"], b.get("purpose", ""), b.get("affects", ""),
                                b.get("local", b.get("d49", "")), run_info.get("AI model", "")])
        _sheet(wb, "Impact (AI-drafted)",
               ["File", "Title", "Purpose", "Who it affects", "Local impact", "Model"],
               ai_rows, widths=[16, 46, 40, 40, 44, 22], fill=AI_FILL, font=AI_FONT)

    sp_rows = []
    for r in rows:
        for s in r.get("_sponsor_objs", []):
            seq = s.get("MatterSponsorSequence", "")
            sp_rows.append([r["File"], r["Type"], s.get("MatterSponsorName", ""),
                            seq, "Prime" if seq == 0 else "Co-sponsor", r["MatterId"]])
    _sheet(wb, "Sponsors", ["File", "Type", "Sponsor", "Seq", "Role", "MatterId"], sp_rows,
           widths=[16, 15, 28, 6, 12, 10])

    h_rows = []
    for r in rows:
        seen_h = set()
        for h in sorted(histories_map.get(r["MatterId"], []), key=lambda x: x.get("MatterHistoryActionDate") or ""):
            d = h.get("MatterHistoryActionDate")
            a = (h.get("MatterHistoryActionName") or "").strip()
            b = h.get("MatterHistoryActionBodyName") or ""
            if not d or not a:
                continue
            k = (d, a, b)
            if k in seen_h:  # duplicate data-entry rows happen; drop them
                continue
            seen_h.add(k)
            h_rows.append([r["File"], _date(d), a, b, h.get("MatterHistoryPassedFlagName", "")])
    _sheet(wb, "History", ["File", "Action Date", "Action", "Action By", "Result"], h_rows,
           widths=[16, 12, 40, 30, 12])

    a_rows = []
    for r in rows:
        for a in attach_map.get(r["MatterId"], []):
            a_rows.append([r["File"], a.get("MatterAttachmentName", ""), a.get("MatterAttachmentHyperlink", "")])
    _sheet(wb, "Attachments", ["File", "Attachment", "Link"], a_rows, widths=[16, 40, 70])

    if text_map:
        t_rows = [[r["File"], r["Title"], (text_map.get(r["MatterId"], "") or "")[:30000]]
                  for r in rows if text_map.get(r["MatterId"])]
        _sheet(wb, "Bill Text", ["File", "Title", "Plain Text (truncated 30k)"], t_rows,
               widths=[16, 50, 120])

    hdr = ["File", "Change", "Field", "Was", "Now", "When (last modified UTC)"]
    if changes:
        _sheet(wb, "Changes Since Last Sync", hdr, changes, widths=[16, 14, 16, 36, 36, 20], fill=CHANGE_FILL)
    else:
        msg = run_info.get("Changes placeholder", "(no changes since last sync)")
        _sheet(wb, "Changes Since Last Sync", hdr, [[msg, "", "", "", "", ""]], widths=[16, 14, 16, 36, 36, 20])

    _sheet(wb, "Run Info", ["Field", "Value"], [[k, v] for k, v in run_info.items()], widths=[26, 80])
    wb.move_sheet("Matters", -(len(wb.sheetnames) - 1))
    wb.save(path)
    return path


# ============================================================================
# ORCHESTRATION — shared by CLI and the scheduled runner
# ============================================================================
def assemble(client, ai, snap, old, profile):
    """Run one profile. Returns a bundle dict ready for build_workbook. Does NOT save snapshot."""
    f = profile.get("filter", {})
    clauses = []
    if f.get("file"):  clauses.append(f"MatterFile eq '{f['file']}'")
    if f.get("since"): clauses.append(f"MatterIntroDate ge datetime'{f['since']}'")
    if f.get("until"): clauses.append(f"MatterIntroDate lt datetime'{f['until']}'")
    if f.get("type"):  clauses.append(f"MatterTypeName eq '{f['type']}'")
    odata = " and ".join(clauses) or None

    raw = client.matters(odata)
    raw.sort(key=lambda m: m.get("MatterIntroDate") or "", reverse=True)  # safe in-memory display order
    if profile.get("limit"):
        raw = raw[: profile["limit"]]

    enrich = profile.get("enrich", False)
    enrich_status = set(profile.get("enrich_status") or [])  # empty => enrich all
    want_text = profile.get("text", False)
    workers = profile.get("workers", 4)
    sponsor_filter = (f.get("sponsor") or "").lower()

    sponsors_map, histories_map, attach_map, text_map = {}, {}, {}, {}

    def _pool(items, fn, label):
        total = len(items)
        done = 0
        with ThreadPoolExecutor(max_workers=workers) as ex:
            for fut in as_completed([ex.submit(fn, m) for m in items]):
                fut.result()
                done += 1
                if total > 40 and done % 25 == 0:
                    print(f"   {label}: {done}/{total}")

    # Phase A — sponsor scan (cheap: 1 call/matter) only when filtering by sponsor
    scan_failures = [0]
    if sponsor_filter:
        print(f"Scanning {len(raw)} bills for sponsor '{f['sponsor']}' — this is the slow part, please wait...")

        def _scan(m):
            try:
                sponsors_map[m["MatterId"]] = client.sponsors(m["MatterId"]) or []
            except Exception:
                sponsors_map[m["MatterId"]] = []
                scan_failures[0] += 1
        _pool(raw, _scan, "scanned")
        raw = [m for m in raw
               if any(sponsor_filter in (s.get("MatterSponsorName") or "").lower()
                      for s in current_sponsors(m, sponsors_map.get(m["MatterId"], [])))]
        print(f"   matched {len(raw)} bills sponsored by '{f['sponsor']}' "
              f"({scan_failures[0]} sponsor lookups failed)")

    # Phase B — heavy detail (histories/attachments/text) only on the matters we keep
    if enrich:
        targets = [m for m in raw if not enrich_status or m.get("MatterStatusName") in enrich_status]
        sponsors_only = profile.get("sponsors_only", False)

        def heavy(m):
            mid = m["MatterId"]
            try:
                if mid not in sponsors_map:
                    sponsors_map[mid] = client.sponsors(mid) or []
                if sponsors_only:
                    return
                histories_map[mid] = client.histories(mid)
                attach_map[mid] = client.attachments(mid)
                if want_text:
                    tx = client.text_plain(mid, m.get("MatterVersion"))
                    if tx:
                        text_map[mid] = tx
            except Exception:
                scan_failures[0] += 1
                sponsors_map.setdefault(mid, [])

        if targets:
            print(f"Pulling {'sponsors' if sponsors_only else 'full details'} for {len(targets)} bills...")
            _pool(targets, heavy, "details")

    rows = []
    for m in raw:
        mid = m["MatterId"]
        rows.append(normalize_matter(
            m,
            sponsors_map.get(mid) if mid in sponsors_map else None,
            histories_map.get(mid),
            attach_map.get(mid) if mid in attach_map else None))

    impact_mode = profile.get("impact", "none")
    ai_map = {}
    ai_drafted = ai_cached = 0
    if impact_mode == "ai" and ai is not None:
        cap = profile.get("ai_max_bills", 0)
        consider = rows[:cap] if cap else rows
        for r in consider:
            mid = r["MatterId"]
            th = text_hash(r.get("Title", ""), text_map.get(mid, ""))
            cached = snap.ai_get(mid, th, ai.model)
            if cached:
                ai_map[mid] = cached; ai_cached += 1; continue
            try:
                b = ai.draft(r, text_map.get(mid, ""))
                snap.ai_put(mid, th, ai.model, b); ai_map[mid] = b; ai_drafted += 1
            except Exception as e:
                ai_map[mid] = {"purpose": f"[AI error: {e}]", "affects": "", "local": ""}

    changes = diff(old, rows) if old else []
    run_info = {
        "Profile": profile.get("name", "(cli)"),
        "Run timestamp (UTC)": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        "API base": API_BASE, "Filter": odata or "(none)",
        "Sponsor filter": f.get("sponsor", "(none)"),
        "Enrich": enrich, "Enrich status gate": ", ".join(enrich_status) or "(all)",
        "Full text pulled": bool(text_map),
        "Impact mode": impact_mode, "AI model": (ai.model if (impact_mode == "ai" and ai) else "(n/a)"),
        "AI drafted this run": ai_drafted, "AI reused from cache": ai_cached,
        "Matters in workbook": len(rows), "Prior snapshot matters": len(old),
        "Changes detected": len(changes), "Token used": bool(getattr(client, "token", None)),
        "Sponsor scan failures": scan_failures[0],
    }
    return {"rows": rows, "sponsors_map": sponsors_map, "histories_map": histories_map,
            "attach_map": attach_map, "text_map": text_map, "ai_map": ai_map,
            "changes": changes, "run_info": run_info, "impact_mode": impact_mode,
            "_filter": odata, "_raw_count": len(raw)}


def main():
    ap = argparse.ArgumentParser(description="Sync NYC Council legislation 1:1 from Legistar (single profile).")
    ap.add_argument("--file"); ap.add_argument("--since"); ap.add_argument("--until")
    ap.add_argument("--type"); ap.add_argument("--sponsor")
    ap.add_argument("--enrich", action="store_true")
    ap.add_argument("--enrich-status", help="Comma list, e.g. Committee — enrich only these statuses")
    ap.add_argument("--text", action="store_true")
    ap.add_argument("--impact", choices=["none", "keyword", "ai"], default="none")
    ap.add_argument("--ai-model", default=os.environ.get("LEGISTAR_AI_MODEL", "claude-haiku-4-5-20251001"))
    ap.add_argument("--limit", type=int); ap.add_argument("--workers", type=int, default=6)
    ap.add_argument("--state", default="legistar_state.db"); ap.add_argument("--out", default="legislation.xlsx")
    a = ap.parse_args()

    client = LegistarClient()
    snap = Snapshot(a.state)
    old = snap.load()
    ai = AIImpact(a.ai_model) if a.impact == "ai" else None
    profile = {"name": "cli", "filter": {k: getattr(a, k) for k in ("file", "since", "until", "type", "sponsor") if getattr(a, k)},
               "enrich": a.enrich, "enrich_status": (a.enrich_status.split(",") if a.enrich_status else None),
               "text": a.text, "impact": a.impact, "limit": a.limit, "workers": a.workers}
    print(f"[*] {profile['filter'] or '(no filter)'}  impact={a.impact}")
    bundle = assemble(client, ai, snap, old, profile)
    snap.save(bundle["rows"])
    build_workbook(bundle, a.out)
    ri = bundle["run_info"]
    print(f"[*] {ri['Matters in workbook']} matters, {ri['Changes detected']} changes, "
          f"AI drafted {ri['AI drafted this run']} / reused {ri['AI reused from cache']} -> {a.out}")


import collections
def _status_group(s):
    ALIVE = {"Committee", "Laid Over in Committee"}
    PASSED = {"Adopted", "Enacted", "Enacted (Mayor's Desk for Signature)"}
    return "Alive" if s in ALIVE else ("Passed" if s in PASSED else "Dead/Filed")

def overview_stats(rows):
    g = collections.Counter(_status_group(r.get("Status", "")) for r in rows)
    return {"total": len(rows),
            "alive": g.get("Alive", 0), "passed": g.get("Passed", 0), "dead": g.get("Dead/Filed", 0)}

def pillar_counts(rows):
    c = collections.Counter()
    for r in rows:
        for t in (r.get("Topic Tags") or r.get("Topic tags") or "").split("; "):
            if t: c[t] += 1
    return dict(c)

def status_counts(rows):
    return dict(collections.Counter(_status_group(r.get("Status", "")) for r in rows))

def coalition_counts(rows, member, top=15):
    c = collections.Counter()
    for r in rows:
        names = r.get("_sponsor_names", [])
        if any(member in n for n in names):
            for n in names:
                if member not in n: c[n] += 1
    return dict(c.most_common(top))

import re as _re
def parse_file(fileno):
    """'Int 0220-2026' -> ('Int', 220, '2026'). Returns (prefix, number_or_None, year_or_None)."""
    m = _re.match(r"\s*([A-Za-z\.]+)\s*0*(\d+)\s*-\s*(\d{4})", fileno or "")
    if m:
        return m.group(1), int(m.group(2)), m.group(3)
    return (fileno or ""), None, None

def _blob(r):
    parts = [str(r.get(k, "") or "") for k in
             ("File", "Title", "Name", "Type", "Status", "Committee/Body", "Prime Sponsor",
              "Topic Tags")]
    parts += list(r.get("_sponsor_names", []))
    return " ".join(parts).lower()

def matches_search(r, query):
    """Smart, simple search. Pure number -> match bill number across ALL types.
    Otherwise every word must appear somewhere (title, sponsors, committee, type, etc.)."""
    if not query or not query.strip():
        return True
    q = query.strip().lower()
    prefix, num, year = parse_file(r.get("File", ""))
    if q.isdigit():
        qn = int(q)
        if num is not None and num == qn:
            return True
        if q in (r.get("File", "") or "").lower():
            return True
        if year and q in year:
            return True
        return False
    return all(tok in _blob(r) for tok in q.split())

def type_counts(rows):
    return dict(collections.Counter(r.get("Type", "") for r in rows if r.get("Type")))

def overview_general(rows):
    g = collections.Counter(_status_group(r.get("Status", "")) for r in rows)
    return {"total": len(rows), "alive": g.get("Alive", 0),
            "passed": g.get("Passed", 0), "dead": g.get("Dead/Filed", 0)}

def member_bills(rows, member):
    m = (member or "").lower().strip()
    if not m:
        return []
    return [r for r in rows if any(m in (n or "").lower() for n in r.get("_sponsor_names", []))
            or m in (r.get("Prime Sponsor", "") or "").lower()]

def member_prime_count(rows, member):
    m = (member or "").lower().strip()
    return sum(1 for r in rows if m in (r.get("Prime Sponsor", "") or "").lower())

def dossier_stats(mb, member):
    m = (member or "").lower()
    prime = member_prime_count(mb, member)
    st_ = overview_general(mb)
    last = member.split()[-1] if member.split() else member
    coal = coalition_counts(mb, member=last)
    prime_titles = [r.get("Title", "") for r in mb if m in (r.get("Prime Sponsor", "") or "").lower()][:8]
    return {"member": member, "bills_on": len(mb), "as_prime": prime, "as_cosponsor": len(mb) - prime,
            "by_topic": pillar_counts(mb),
            "by_status": {"alive": st_["alive"], "passed": st_["passed"], "dead": st_["dead"]},
            "top_coalition": dict(list(coal.items())[:8]),
            "example_prime_bills": prime_titles}


def coalition_edges(rows, top_members=28, min_weight=2):
    """Build a co-sponsorship network from loaded rows.
    Nodes = members (size by # bills sponsored); edges = # bills two members co-sponsored."""
    import collections, itertools
    deg = collections.Counter()
    pair = collections.Counter()
    for r in rows:
        names = sorted({n.strip() for n in (r.get("_sponsor_names") or []) if n and n.strip()})
        for n in names:
            deg[n] += 1
        for a, b in itertools.combinations(names, 2):
            pair[(a, b)] += 1
    top = {m for m, _ in deg.most_common(top_members)}
    nodes = [{"id": m, "label": m, "value": deg[m]} for m in top]
    edges = [{"from": a, "to": b, "value": w}
             for (a, b), w in pair.items() if a in top and b in top and w >= min_weight]
    return nodes, edges


def coalition_matrix(rows, top_members=18):
    """Symmetric co-sponsorship counts among the most active members.
    Returns (members_ordered_by_activity, degree_dict, pair_count_dict)."""
    import collections, itertools
    deg = collections.Counter()
    pair = collections.Counter()
    for r in rows:
        names = sorted({n.strip() for n in (r.get("_sponsor_names") or []) if n and n.strip()})
        for n in names:
            deg[n] += 1
        for a, b in itertools.combinations(names, 2):
            pair[(a, b)] += 1
    members = [m for m, _ in deg.most_common(top_members)]
    return members, dict(deg), dict(pair)

import streamlit as st
import pandas as pd
import datetime as _dt

# v2 feature modules (self-contained; safe to import here — no side effects)
import llm as _llm
import briefing as _brief
import policylab as _lab
import people as _people
import packet as _packet
import store as _store
import messaging as _msg
import analysis as _analysis
import citydata as _city
import memory as _memory
import retrieval as _retrieval
try:
    from sources import media as _media
except Exception:
    _media = None


@st.cache_resource(show_spinner=False)
def _mem():
    """One persistent Memory for the deployment (survives reruns)."""
    return _memory.Memory()


def _get_index(rows):
    """Build/reuse a TF-IDF search index for the loaded rows (cached in session)."""
    key = (st.session_state.get("loaded_year", ""), len(rows))
    cur = st.session_state.get("_index_key")
    if cur != key or "_search_index" not in st.session_state:
        st.session_state["_search_index"] = _retrieval.Index.build(rows, _retrieval.bill_text)
        st.session_state["_index_key"] = key
    return st.session_state["_search_index"]
try:
    from sources import nystate as _nys, congress as _cong
except Exception:  # keep the app up even if a source module has an issue
    _nys = _cong = None
try:
    from sources import housevotes as _housevotes
except Exception:
    _housevotes = None
try:
    from sources import opendata as _od
except Exception:
    _od = None

st.set_page_config(page_title="NYC Legislative Intelligence", layout="wide", initial_sidebar_state="collapsed")
NYC_TOKEN = "Uvxb0j9syjm3aI8h46DhQvnX5skN4aSUL0x_Ee3ty9M.ew0KICAiVmVyc2lvbiI6IDEsDQogICJOYW1lIjogIk5ZQyByZWFkIHRva2VuIDIwMTcxMDI2IiwNCiAgIkRhdGUiOiAiMjAxNy0xMC0yNlQxNjoyNjo1Mi42ODM0MDYtMDU6MDAiLA0KICAiV3JpdGUiOiBmYWxzZQ0KfQ"

def year_window(year):
    if year == "2024–present":
        return "2024-01-01", None
    y = int(year)
    return f"{y}-01-01", f"{y + 1}-01-01"

st.markdown("""
<style>
/* ===== Light, cohesive design system ===== */
:root { --bg:#f5f7fb; --bg2:#eef2f9; --surf:#ffffff; --surf2:#f3f6fc; --line:#e2e8f2;
        --ink:#1a2537; --mut:#5b6b86; --blue:#1d4ed8; --blue2:#2563eb; --cyan:#0891b2;
        --teal:#0d9488; --green:#059669; --shadow:0 2px 10px rgba(24,45,90,.06);
        --shadow2:0 6px 22px rgba(24,45,90,.10); }
[data-testid="stSidebar"] { display:none !important; }
[data-testid="stSidebarCollapsedControl"] { display:none !important; }
.stApp { background:
  radial-gradient(1100px 520px at 88% -12%, #e7eefb 0%, rgba(231,238,251,0) 60%),
  radial-gradient(900px 440px at -8% 4%, #eaf1ff 0%, rgba(234,241,255,0) 55%), var(--bg) fixed;
  color: var(--ink); }
.block-container { padding-top: 1.0rem; max-width: 1400px; }
body, .stMarkdown, p, span, label, li { color: var(--ink); }
.appbar { background: linear-gradient(115deg,#12275a 0%,#1e40af 55%,#2563eb 130%);
  color:#fff; border-radius:18px; padding:20px 26px; margin-bottom:16px;
  border:1px solid #1b3a86; box-shadow:0 10px 30px rgba(29,78,216,.22); position:relative; overflow:hidden; }
.appbar:after { content:""; position:absolute; right:-30px; top:-70px; width:240px; height:240px;
  background:radial-gradient(circle, rgba(255,255,255,.16) 0%, rgba(255,255,255,0) 70%); }
.appbar-title { font-size:1.55rem; font-weight:800; letter-spacing:.2px; color:#fff; }
.appbar-sub { opacity:.92; font-size:.9rem; margin-top:4px; color:#dbe6ff; }
.livepill { position:absolute; top:20px; right:24px; background:rgba(255,255,255,.16);
  color:#eafff4; font-weight:700; font-size:.7rem; padding:4px 11px; border-radius:999px; letter-spacing:.6px;
  border:1px solid rgba(255,255,255,.35); }
div[data-testid="stMetric"] { background:var(--surf); border:1px solid var(--line); border-radius:14px;
  padding:14px 16px; box-shadow:var(--shadow); }
div[data-testid="stMetricValue"] { color:var(--blue); font-weight:800; }
div[data-testid="stMetricLabel"] { color:var(--mut); font-weight:600; }
.stTabs [data-baseweb="tab-list"] { gap:6px; flex-wrap:wrap; border-bottom:1px solid var(--line); }
.stTabs [data-baseweb="tab"] { background:var(--surf); border:1px solid var(--line); border-bottom:none;
  border-radius:11px 11px 0 0; padding:7px 14px; font-weight:600; color:var(--mut); }
.stTabs [aria-selected="true"] { background:linear-gradient(180deg,#2563eb,#1d4ed8) !important; color:#fff !important;
  border-color:#1d4ed8; }
.stButton>button { background:linear-gradient(180deg,#2563eb,#1d4ed8); color:#fff; border:1px solid #1d4ed8;
  border-radius:10px; font-weight:700; padding:.5rem 1.1rem; box-shadow:0 3px 10px rgba(37,99,235,.22); }
.stButton>button:hover { background:linear-gradient(180deg,#1d4ed8,#1e3a8a); color:#fff; }
.stDownloadButton>button { background:linear-gradient(180deg,#0d9488,#0f766e); color:#fff; border:1px solid #0f766e;
  border-radius:10px; font-weight:700; }
[data-testid="stExpander"] { border:1px solid var(--line); border-radius:14px; background:var(--surf);
  box-shadow:var(--shadow); }
[data-testid="stExpander"] summary { font-weight:700; color:var(--ink); }
[data-testid="stDataFrame"] { border:1px solid var(--line); border-radius:12px; }
[data-baseweb="select"]>div, .stTextInput input, .stNumberInput input, .stTextArea textarea {
  background:var(--surf) !important; color:var(--ink) !important; border-color:var(--line) !important; }
a { color:var(--blue) !important; }
h1,h2,h3 { color:#132444; }
div[data-testid="stAlert"] { border-radius:12px; border:1px solid var(--line); }
hr { border-color:var(--line); }

/* ---- component system ---- */
.stTabs .stTabs [data-baseweb="tab"] { padding:5px 11px; font-size:.86rem; }
.kicker { text-transform:uppercase; letter-spacing:.14em; font-size:.7rem; font-weight:800; color:var(--mut); }
.hero { background:linear-gradient(120deg,#12275a 0%,#1e40af 60%,#2563eb 130%);
  border-radius:18px; padding:22px 26px; margin-bottom:16px; position:relative; overflow:hidden;
  box-shadow:0 10px 30px rgba(29,78,216,.22); }
.hero h1 { font-size:1.7rem; font-weight:800; margin:0 0 4px; color:#fff; }
.hero p { color:#dbe6ff; margin:0; }
.hero:after { content:""; position:absolute; right:-40px; top:-80px; width:280px; height:280px;
  background:radial-gradient(circle, rgba(255,255,255,.16) 0%, rgba(255,255,255,0) 70%); }
.badge { display:inline-block; padding:2px 10px; border-radius:999px; font-size:.72rem; font-weight:700;
  letter-spacing:.02em; border:1px solid transparent; }
.b-nyc   { background:#e6effe; color:#1746b0; border-color:#c5dafb; }
.b-state { background:#f0e9fe; color:#6b34c9; border-color:#dbccfa; }
.b-fed   { background:#fde9e9; color:#c02626; border-color:#f7cdcd; }
.b-muted { background:var(--surf2); color:var(--mut); border-color:var(--line); }
.b-green { background:#e3f6ee; color:#08794a; border-color:#c4ead8; }
.card { background:var(--surf); border:1px solid var(--line); border-radius:14px;
  padding:14px 16px; margin-bottom:10px; box-shadow:var(--shadow); }
.card h4 { margin:0 0 4px; color:#132444; font-size:1.02rem; }
.card .meta { color:var(--mut); font-size:.82rem; }
.pcard { border-left:4px solid var(--blue); }
.pcard.state { border-left-color:#7c3aed; } .pcard.fed { border-left-color:#ef4444; }
.brief { background:#f7faff; border:1px solid var(--line); border-left:4px solid var(--blue);
  border-radius:12px; padding:6px 20px 14px; box-shadow:var(--shadow); }
.brief h2 { color:#12275a; font-size:1.15rem; margin-top:14px; }
.brief h1 { color:#12275a; }
.chip { display:inline-block; background:var(--surf2); border:1px solid var(--line); color:var(--mut);
  border-radius:8px; padding:2px 9px; font-size:.75rem; margin:2px 4px 2px 0; }
.stars { color:#e0a400; letter-spacing:2px; }
.memberrow { display:flex; align-items:center; gap:12px; }
.memberrow .info h4 { margin:0; } .memberrow .info .meta { color:var(--mut); font-size:.82rem; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="appbar">
  <div class="appbar-title">🗽 NYC Legislative Intelligence</div>
  <div class="appbar-sub"><b>City Hall first</b> — the Council, its members &amp; districts — with Albany &amp; Washington alongside.
  &nbsp;<b>“Bulletpoints for Bureaucrats”</b> briefings built for the desk.</div>
  <span class="livepill">● LIVE</span>
</div>
""", unsafe_allow_html=True)

YEAR_OPTS = ["2026", "2025", "2024", "2023", "2022", "2024–present"]
SCOPES = ["All legislation (browse list)", "By a Council Member", "One specific bill"]

@st.cache_resource
def _client():
    return LegistarClient(token=NYC_TOKEN, pause=0.2)

@st.cache_data(ttl=1800, show_spinner=False)
def fetch_detail(mid):
    c = _client(); out = {"sponsors": [], "histories": [], "attachments": [], "text": ""}
    for k, fn in [("sponsors", c.sponsors), ("histories", c.histories), ("attachments", c.attachments)]:
        try: out[k] = fn(mid)
        except Exception: pass
    try: out["text"] = c.text_plain(mid, None)
    except Exception: pass
    return out

@st.cache_data(ttl=1800, show_spinner=False)
def fetch_votes(mid):
    c = _client()
    try:
        hi = c.histories(mid)
    except Exception:
        return []
    return matter_vote_events(c, hi)

@st.cache_data(ttl=1800, show_spinner=False)
def fetch_hearings(frm, to):
    c = _client()
    flt = f"EventDate ge datetime'{frm}' and EventDate lt datetime'{to}'"
    r = [normalize_event(e) for e in c.events(flt)]
    r.sort(key=lambda x: x["_sortdate"])
    return r

@st.cache_data(ttl=1800, show_spinner=False)
def fetch_agenda(event_id):
    items = _client().event_items(event_id); out = []
    for it in items:
        if not (it.get("EventItemTitle") or it.get("EventItemMatterFile")):
            continue
        out.append({"Seq": it.get("EventItemAgendaSequence"), "File": it.get("EventItemMatterFile", "") or "",
                    "Item": (it.get("EventItemTitle") or "").strip(),
                    "Action": it.get("EventItemActionName", "") or "", "Result": it.get("EventItemPassedFlagName", "") or ""})
    out.sort(key=lambda x: (x["Seq"] is None, x["Seq"] or 0))
    return out

@st.cache_data(ttl=86400, show_spinner=False)
def get_directory():
    try: return _client().council_members()
    except Exception: return []

@st.cache_data(ttl=86400, show_spinner=False)
def get_committees():
    try: return _client().committees()
    except Exception: return []

# --- top control panel (replaces the old sidebar) ---
_dir = get_directory()
_pre_bundle = st.session_state.get("bundle")
with st.expander("⚙️  Data controls — choose what to load, then press Load", expanded=(_pre_bundle is None)):
    a1, a2, a3 = st.columns([1, 1.5, 1.1])
    year = a1.selectbox("Year", YEAR_OPTS, index=0)
    scope = a2.selectbox("Scope", SCOPES)
    include_sponsors = a3.checkbox("Include sponsors (slower)", value=False,
        help="On = pull who signed each bill so the list is searchable by member. Off loads much faster.")
    if scope == "By a Council Member":
        member_name = st.selectbox("Council Member", _dir, key="cp_member") if _dir else st.text_input("Council Member (last name)", "", key="cp_member_txt")
    else:
        member_name = ""
    if scope == "One specific bill":
        _loaded_files = sorted({r["File"] for r in (_pre_bundle or {}).get("rows", []) if r.get("File")})
        if _loaded_files:
            bill_number = st.selectbox("Pick a bill", _loaded_files, key="cp_bill")
        else:
            bill_number = st.text_input("Bill number", "Int 0220-2026", key="cp_bill_txt")
            st.caption("Tip: load **All legislation** once — then this becomes a dropdown of every bill.")
    else:
        bill_number = ""
    b1, b2, b3 = st.columns([1.3, 1.3, 1])
    add_ai = b1.checkbox("Add AI impact bullets", value=False)
    anthropic_key = b2.text_input("Anthropic key (optional)", "", type="password")
    b3.write(""); b3.write("")
    load = b3.button("⟳  Load data", type="primary")
    st.caption("First load of a full year reads live from NYC's servers (seconds–minutes); cached and instant after. "
               "Single-bill and member scopes are fastest.")

def _tag_rows(rows, text_map=None):
    for r in rows:
        t = keyword_tags(r, ((text_map or {}).get(r["MatterId"], "") or "")[:4000])
        r["Topic tags"] = t["Topic Tags"]; r["Boroughs named"] = t["Boroughs Named"]; r["Topic Tags"] = t["Topic Tags"]
    return rows

@st.cache_data(ttl=1800, show_spinner=False)
def build_member_dossier(member, year):
    since, until = year_window(year)
    flt = {"since": since, "sponsor": member}
    if until: flt["until"] = until
    profile = {"name": "dossier", "filter": flt, "enrich": True, "sponsors_only": True, "text": False, "workers": 10, "impact": "keyword"}
    snap = Snapshot("/tmp/legistar_state.db"); old = snap.load()
    bundle = assemble(_client(), None, snap, old, profile)
    _tag_rows(bundle["rows"])
    return {"member": member, "rows": bundle["rows"], "stats": dossier_stats(bundle["rows"], member),
            "scanned": bundle.get("_raw_count"), "scan_failures": bundle["run_info"].get("Sponsor scan failures", 0)}

@st.cache_data(ttl=86400, show_spinner=False)
def make_dossier_ai(member, stats, key):
    return AIImpact("claude-haiku-4-5-20251001", api_key=key).dossier(member, stats)

@st.cache_data(ttl=1800, show_spinner=False)
def run_pull(scope, year, include_sponsors, bill_number, member_name, add_ai, anthropic_key):
    since, until = year_window(year)
    profile = {"name": "web", "filter": {}, "enrich": True, "text": True, "workers": 4}
    if scope == "By a Council Member":
        flt = {"since": since, "sponsor": member_name.strip()}
        if until: flt["until"] = until
        profile["filter"] = flt; profile["text"] = False
    elif scope == "All legislation (browse list)":
        flt = {"since": since}
        if until: flt["until"] = until
        profile["filter"] = flt; profile["text"] = False
        if include_sponsors:
            profile["enrich"] = True; profile["sponsors_only"] = True; profile["workers"] = 6
        else:
            profile["enrich"] = False
    else:
        profile["filter"] = {"file": bill_number.strip()}
    profile["impact"] = "ai" if (add_ai and anthropic_key.strip()) else "keyword"
    if add_ai and anthropic_key.strip():
        os.environ["ANTHROPIC_API_KEY"] = anthropic_key.strip()
    client = LegistarClient(token=NYC_TOKEN, pause=0.2)
    snap = Snapshot("/tmp/legistar_state.db"); old = snap.load()
    ai = AIImpact("claude-haiku-4-5-20251001") if profile["impact"] == "ai" else None
    bundle = assemble(client, ai, snap, old, profile)
    if scope == "All legislation (browse list)" and not bundle["rows"] and until:
        profile["filter"] = {"since": since}
        bundle = assemble(client, ai, snap, old, profile)
    _tag_rows(bundle["rows"], bundle["text_map"])
    snap.save(bundle["rows"]); build_workbook(bundle, "/tmp/legislation.xlsx")
    return bundle

with st.expander("🔧 Connection test — open this if a year won't load"):
    st.caption("Runs the exact queries the loader uses and shows what NYC Legistar returns, so we can see the problem.")
    if st.button("Run connection test"):
        c = _client()
        for label, flt in [("2026", "MatterIntroDate ge datetime'2026-01-01' and MatterIntroDate lt datetime'2027-01-01'"),
                           ("2025", "MatterIntroDate ge datetime'2025-01-01' and MatterIntroDate lt datetime'2026-01-01'"),
                           ("no filter (first page)", None)]:
            try:
                res = c.matters(flt)
                st.write(f"**{label}** → {len(res)} matters. Sample: " + ", ".join(m.get('MatterFile','?') for m in res[:6]))
            except Exception as e:
                st.error(f"**{label}** query failed: {type(e).__name__}: {e}")

if load:
    try:
        with st.spinner("Working... pulling from Legistar. Big years can take a moment."):
            _b = run_pull(scope, year, include_sponsors, bill_number, member_name, add_ai, anthropic_key)
        if not _b["rows"]:
            run_pull.clear()  # never trap a transient/empty result in the cache
        st.session_state["bundle"] = _b
        st.session_state["loaded_year"] = year
    except requests.exceptions.HTTPError as e:
        st.error(f"NYC API returned HTTP {getattr(e.response,'status_code','?')}: {(getattr(e.response,'text','') or '')[:400]}")
    except Exception as e:
        st.error(f"{type(e).__name__}: {e}")

bundle = st.session_state.get("bundle")
rows = bundle["rows"] if bundle else []
loaded_year = st.session_state.get("loaded_year", "")
if not bundle:
    st.info("**Open the ⚙️ Data controls panel above, pick a Year and Scope, then press _Load data_.**  \n"
            "Start with **All legislation** + **2026** to load every Introduction, Resolution, and Land Use item "
            "for the year — then use the **Legislation list** tab to search by number or word.")
elif not rows:
    st.warning(f"This load returned **0 bills**.  Filter used: `{bundle.get('_filter')}` · raw matches from "
               f"Legistar: **{bundle.get('_raw_count')}**.  Try another year, or clear filters in the controls panel.")
elif load:
    st.success(f"Loaded **{len(rows)}** items for **{loaded_year}**. Use the tabs below.")

# Grouped, two-level navigation: 7 clean sections, each with focused sub-tabs.
# (Streamlit tab containers carry their own path, so the `with t_x:` blocks
#  below can stay where they are and still render inside the right section.)
# City-first navigation: City Hall leads; State & Federal is a clearly separate section.
sec_home, sec_city, sec_leg, sec_people, sec_brief, sec_politics, sec_levels, sec_ask, sec_about = st.tabs(
    ["🏛️ Command Center", "🏙️ City Hall", "📜 Legislation", "👥 People & Coalitions",
     "📰 Briefings & Ideas", "📣 Politics & Messaging", "🇺🇸 State & Federal", "💬 Ask", "ℹ️ About"])
t_home, t_ask, t_about = sec_home, sec_ask, sec_about
with sec_city:
    t_officials, t_council, t_distprofile, t_reps = st.tabs(
        ["🏛️ City Officials", "🧑‍🤝‍🧑 Council Members", "📍 District Profile", "🏠 Find my reps"])
with sec_politics:
    t_warroom, t_statement, t_rapid, t_influence = st.tabs(
        ["🎯 Issue War Room", "📝 Statement Studio", "⚡ Rapid Response", "🧭 Influence Map"])
with sec_leg:
    t_list, t_detail, t_hear, t_changes, t_over = st.tabs(
        ["📋 Legislation list", "📄 Bill detail", "📅 Hearings", "🔔 What changed", "📊 Overview"])
with sec_levels:
    t_gov, t_votes, t_activity, t_dir, t_elect = st.tabs(
        ["🏙️ State & Federal bills", "🗳️ Votes & decisions", "🔔 Activity (all levels)",
         "👤 Who governs NYC", "🗳️ Elections & terms"])
with sec_people:
    t_members, t_wiki, t_grid, t_profile, t_dossier, t_compare, t_net, t_map = st.tabs(
        ["👤 Members", "📖 CM Wiki", "📊 Policy Grid", "🪪 Deep profile", "📕 Dossier",
         "⚖️ Compare", "🤝 Coalitions", "🗺️ District map"])
with sec_brief:
    t_brief, t_packet, t_lab, t_memory = st.tabs(
        ["📰 Briefing Studio", "📦 District Packet", "💡 Policy Lab", "🧠 Knowledge & Memory"])

def need_data():
    st.info("Load data from the ⚙️ controls panel above first (this tab uses that data).")

# ---------------- LEGISLATION LIST (Legistar-style master list + search) ----------------
with t_list:
    if not bundle:
        need_data()
    else:
        st.subheader(f"All legislation — {loaded_year}")
        st.caption("This is the full list (every type). Type a number like **220** to find that bill across all "
                   "types, or words like **ferry**, a committee, or a member's name. Leave blank to see everything.")
        q = st.text_input("Search", key="ls_q")
        cc = st.columns(4)
        types = sorted({r["Type"] for r in rows if r["Type"]})
        pick_type = cc[0].multiselect("Type (Introduction / Resolution / Land Use…)", types, key="ls_type")
        statuses = sorted({r["Status"] for r in rows if r["Status"]})
        pick_status = cc[1].multiselect("Status", statuses, key="ls_status")
        topics = sorted({p for r in rows for p in (r.get("Topic tags") or "").split("; ") if p})
        pick_topic = cc[2].multiselect("Policy topic", topics, key="ls_topic")
        pick_bor = cc[3].multiselect("Borough named", ["Manhattan", "Brooklyn", "Queens", "Bronx", "Staten Island"], key="ls_bor")
        sponsor_q = st.text_input("Signed on by (member name contains) — needs 'Include sponsors' or a member scope", key="ls_sponsor")

        with st.expander("💾 Saved searches"):
            saved = st.session_state.setdefault("saved_searches", {})
            sc = st.columns([3, 1])
            nm = sc[0].text_input("Name this search", key="save_name")
            if sc[1].button("Save current", key="save_btn"):
                if nm.strip():
                    saved[nm.strip()] = {"q": q, "type": pick_type, "status": pick_status,
                                         "topic": pick_topic, "bor": pick_bor, "sponsor": sponsor_q}
                    st.success(f"Saved '{nm.strip()}'.")
                else:
                    st.warning("Give the search a name first.")
            for nm2 in list(saved.keys()):
                s = saved[nm2]
                bits = []
                if s.get("q"): bits.append(f"\u201C{s['q']}\u201D")
                for lbl, val in [("type", s.get("type")), ("status", s.get("status")),
                                 ("topic", s.get("topic")), ("borough", s.get("bor"))]:
                    if val: bits.append(f"{lbl}: {', '.join(val)}")
                if s.get("sponsor"): bits.append(f"sponsor: {s['sponsor']}")
                rowc = st.columns([3, 1, 1])
                rowc[0].markdown(f"**{nm2}** — {'; '.join(bits) or 'all bills'}")
                if rowc[1].button("Apply", key=f"apply_{nm2}"):
                    st.session_state["ls_q"] = s.get("q", "")
                    st.session_state["ls_type"] = [x for x in s.get("type", []) if x in types]
                    st.session_state["ls_status"] = [x for x in s.get("status", []) if x in statuses]
                    st.session_state["ls_topic"] = [x for x in s.get("topic", []) if x in topics]
                    st.session_state["ls_bor"] = s.get("bor", [])
                    st.session_state["ls_sponsor"] = s.get("sponsor", "")
                    st.rerun()
                if rowc[2].button("Delete", key=f"del_{nm2}"):
                    saved.pop(nm2, None); st.rerun()
            st.caption("Saved searches last while the app is open.")

        f = [r for r in rows if matches_search(r, q)]
        if pick_type:   f = [r for r in f if r["Type"] in pick_type]
        if pick_status: f = [r for r in f if r["Status"] in pick_status]
        if pick_topic:  f = [r for r in f if any(p in (r.get("Topic tags") or "") for p in pick_topic)]
        if pick_bor:    f = [r for r in f if any(b in (r.get("Boroughs named") or "") for b in pick_bor)]
        if sponsor_q.strip():
            sq = sponsor_q.strip().lower()
            f = [r for r in f if any(sq in (n or "").lower() for n in r.get("_sponsor_names", []))
                 or sq in (r.get("Prime Sponsor", "") or "").lower()]
        st.caption(f"Showing {len(f)} of {len(rows)} bills")
        has_sp = any(r.get("_sponsor_names") for r in f)
        disp = []
        for r in f:
            d = {k: v for k, v in r.items() if not k.startswith("_") and k != "Topic Tags"}
            if has_sp: d["All sponsors"] = "; ".join(r.get("_sponsor_names", []))
            disp.append(d)
        if disp:
            st.dataframe(pd.DataFrame(disp), use_container_width=True, height=520,
                column_config={"Web Link": st.column_config.LinkColumn("Legistar", display_text="Open")})

        with st.expander("🔬 Generate full analysis for these bills → Excel (uses your key, capped)"):
            st.caption("Runs the deep analysis on the filtered bills above and returns an Excel. Capped to keep "
                       "time and cost sane — filter the list first, then analyze.")
            ncap = st.number_input("How many of the filtered bills (max 15)", 1, 15, 5)
            if st.button("Generate analyses", key="batch_an"):
                if not anthropic_key.strip():
                    st.warning("Add your Anthropic key in the ⚙️ controls panel above.")
                else:
                    sub = f[:int(ncap)]
                    ai_an = AIImpact("claude-haiku-4-5-20251001", api_key=anthropic_key.strip())
                    out = []; prog = st.progress(0.0)
                    for i, rr in enumerate(sub):
                        try:
                            det = fetch_detail(rr["MatterId"])
                            rr2 = dict(rr); rr2["_sponsor_objs"] = current_sponsors({"MatterVersion": None}, det.get("sponsors", []))
                            anx = ai_an.analyze(rr2, det.get("text", ""), build_data_context(rr))
                        except Exception as e:
                            anx = f"(failed: {e})"
                        out.append({"File": rr["File"], "Title": rr["Title"], "Analysis": anx})
                        prog.progress((i + 1) / len(sub))
                    from openpyxl import Workbook
                    wb = Workbook(); ws = wb.active; ws.title = "Analysis"
                    ws.append(["File", "Title", "Analysis"])
                    for o in out:
                        ws.append([o["File"], o["Title"], o["Analysis"]])
                    wb.save("/tmp/analysis.xlsx"); st.session_state["analysis_xlsx"] = True
                    st.success(f"Analyzed {len(out)} bills.")
            if st.session_state.get("analysis_xlsx"):
                with open("/tmp/analysis.xlsx", "rb") as fh:
                    st.download_button("⬇️ Download analyses (Excel)", fh.read(), "bill_analyses.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- ASK (AI chat over loaded data) ----------------
with t_ask:
    st.subheader("💬 Ask the legislation")
    if not bundle:
        need_data()
    elif not anthropic_key.strip():
        st.info("Add your Anthropic key in the ⚙️ controls panel to use the AI chat.")
    else:
        st.caption("Ask anything about NYC legislation **and** city government — bills, the legislative process, the "
                   "City Charter, the Administrative Code, ULURP, the budget process, what an agency does, or *find me "
                   "the code section on…*. Bill facts come from your loaded data; law/process answers can use web search.")
        allow_web = st.checkbox("🌐 Allow web search (find codes, ordinances, current info)", value=True)

        def _resolve_ref(ref, rws):
            ref = (ref or "").strip()
            for r in rws:
                if (r["File"] or "").lower() == ref.lower():
                    return r
            digits = "".join(ch for ch in ref if ch.isdigit())
            if digits:
                n = int(digits)
                for r in rws:
                    _, num, _ = parse_file(r["File"])
                    if num == n:
                        return r
            for r in rws:
                if ref.lower() and ref.lower() in (r["File"] or "").lower():
                    return r
            return None

        def _apply_filters(rws, fl):
            f = rws
            if fl.get("keyword"): f = [r for r in f if matches_search(r, fl["keyword"])]
            if fl.get("topic"): f = [r for r in f if str(fl["topic"]).lower() in (r.get("Topic tags") or "").lower()]
            if fl.get("type"): f = [r for r in f if str(fl["type"]).lower() in (r.get("Type") or "").lower()]
            if fl.get("status"): f = [r for r in f if str(fl["status"]).lower() in (r.get("Status") or "").lower()]
            if fl.get("borough"): f = [r for r in f if str(fl["borough"]) in (r.get("Boroughs named") or "")]
            if fl.get("sponsor"):
                sq = str(fl["sponsor"]).lower()
                f = [r for r in f if any(sq in (n or "").lower() for n in r.get("_sponsor_names", []))
                     or sq in (r.get("Prime Sponsor", "") or "").lower()]
            if fl.get("min_sponsors"):
                try:
                    n = int(fl["min_sponsors"])
                    f = [r for r in f if isinstance(r.get("Sponsors (#)"), int) and r["Sponsors (#)"] >= n]
                except Exception:
                    pass
            return f

        def _run_agent(q):
            ai = AIImpact("claude-haiku-4-5-20251001", api_key=anthropic_key.strip())
            plan = ai.chat_plan(q) or {}
            evidence = {}
            sponsors_loaded = any(r.get("_sponsor_names") for r in rows)
            for ref in (plan.get("bill_refs") or [])[:4]:
                r = _resolve_ref(ref, rows)
                evidence.setdefault("named_bills", [])
                if not r:
                    evidence["named_bills"].append({"ref": ref, "note": "not in loaded set"}); continue
                item = {"File": r["File"], "Type": r["Type"], "Title": r["Title"], "Status": r["Status"],
                        "Committee": r["Committee/Body"], "Topic": r.get("Topic tags", ""), "Prime": r.get("Prime Sponsor", "")}
                if plan.get("needs_detail"):
                    det = fetch_detail(r["MatterId"])
                    sps = current_sponsors({"MatterVersion": None}, det.get("sponsors", [])) or r.get("_sponsor_objs", [])
                    item["Sponsors"] = [x.get("MatterSponsorName") for x in sps]
                    item["Text excerpt"] = (det.get("text", "") or r.get("Name", ""))[:3000]
                    item["Recent actions"] = [f"{_date(h.get('MatterHistoryActionDate'))}: {(h.get('MatterHistoryActionName') or '').strip()}"
                                              for h in (det.get("histories") or [])[-6:]]
                else:
                    item["Sponsors (#)"] = r.get("Sponsors (#)", "")
                evidence["named_bills"].append(item)
            fl = plan.get("filters") or {}
            if fl or plan.get("intent") == "filter":
                matched = _apply_filters(rows, fl)
                evidence["matched_count"] = len(matched)
                evidence["matched_sample"] = [{"File": r["File"], "Type": r["Type"], "Title": (r["Title"] or "")[:90],
                    "Status": r["Status"], "Sponsors (#)": r.get("Sponsors (#)", ""),
                    "Prime": r.get("Prime Sponsor", ""), "Topic": r.get("Topic tags", "")} for r in matched[:40]]
                if fl.get("min_sponsors") and not sponsors_loaded:
                    evidence["note"] = "Sponsor counts unavailable — load with 'Include sponsors' for accurate min_sponsors filtering."
            if not evidence:
                evidence["overview"] = {"loaded": len(rows), "by_status": status_counts(rows), "by_type": type_counts(rows)}
            return ai.chat_answer(q, evidence, allow_web=allow_web)

        history = st.session_state.setdefault("ask_history", [])
        for msg in history:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])
        with st.form("ask_form", clear_on_submit=True):
            qcol, bcol = st.columns([5, 1])
            q = qcol.text_input("Ask about the loaded bills", label_visibility="collapsed",
                                placeholder="e.g. compare Int 220 to Int 419")
            sent = bcol.form_submit_button("Send")
        if history and st.button("Clear chat", key="ask_clear"):
            st.session_state["ask_history"] = []; st.rerun()
        if sent and q.strip():
            history.append({"role": "user", "content": q})
            with st.spinner("Reading the bills and answering…"):
                try:
                    ans = _run_agent(q)
                except Exception as e:
                    ans = f"Sorry — {type(e).__name__}: {e}"
            history.append({"role": "assistant", "content": ans})
            st.rerun()

# ---------------- HEARINGS ----------------
with t_hear:
    st.subheader("Committee hearings & meetings — schedule")
    cda, cdb, cdc = st.columns([1, 1, 1])
    today = _dt.date.today()
    frm = cda.date_input("From", today); to = cdb.date_input("To", today + _dt.timedelta(days=60))
    cdc.write(""); cdc.write("")
    if cdc.button("Load hearings", type="primary"):
        try:
            with st.spinner("Loading hearings..."):
                st.session_state["hearings"] = fetch_hearings(str(frm), str(to))
        except Exception as e:
            st.error(f"{type(e).__name__}: {e}")
    hev = st.session_state.get("hearings")
    if hev is None:
        st.info("Pick a date range and click **Load hearings** (defaults to the next 60 days).")
    elif not hev:
        st.warning("No hearings found in that range.")
    else:
        bodies = sorted({r["Committee / Body"] for r in hev if r["Committee / Body"]})
        pickb = st.multiselect("Filter by committee", bodies)
        show = [r for r in hev if not pickb or r["Committee / Body"] in pickb]
        st.caption(f"{len(show)} hearings")
        df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_") and k != "EventId"} for r in show])
        st.dataframe(df, use_container_width=True, height=420, column_config={
            "Legistar": st.column_config.LinkColumn("Meeting", display_text="Open"),
            "Agenda": st.column_config.LinkColumn("Agenda", display_text="PDF"),
            "Minutes": st.column_config.LinkColumn("Minutes", display_text="PDF")})
        st.divider()
        st.markdown("**See what's on a hearing's agenda (bills + outcomes):**")
        labels = {f"{r['Date']} — {r['Committee / Body']}": r["EventId"] for r in show}
        if labels:
            picklbl = st.selectbox("Pick a hearing", list(labels.keys()))
            with st.spinner("Loading agenda..."):
                ag = fetch_agenda(labels[picklbl])
            if ag:
                st.dataframe(pd.DataFrame(ag), hide_index=True, use_container_width=True)
            else:
                st.caption("No agenda items posted yet for this hearing.")

# ---------------- BILL DETAIL ----------------
with t_detail:
    if not bundle:
        need_data()
    else:
        pick = st.selectbox("Pick a bill", [r["File"] for r in rows], key="detail")
        r = next(x for x in rows if x["File"] == pick); mid = r["MatterId"]
        _mem().log("view", "bill", r["File"], {"title": (r.get("Title") or "")[:80]})
        sp = r.get("_sponsor_objs", []); hi = bundle["histories_map"].get(mid, [])
        at = bundle["attach_map"].get(mid, []); tx = bundle["text_map"].get(mid, "")
        if not sp and not hi and not at:
            with st.spinner("Loading full details from Legistar..."):
                det = fetch_detail(mid)
            sp = current_sponsors({"MatterVersion": None}, det.get("sponsors", []))
            hi, at, tx = det.get("histories", []), det.get("attachments", []), det.get("text", "")
        st.markdown(f"### {r['File']} — {r['Type']}")
        st.markdown(f"**{r['Title']}**")
        d = st.columns(3)
        d[0].metric("Status", r["Status"] or "-")
        d[1].metric("Sponsors", len(sp) if sp else (r["Sponsors (#)"] if r["Sponsors (#)"] != "" else "-"))
        d[2].write("**Committee**"); d[2].write(r["Committee/Body"] or "-")
        st.markdown(f"[➡ Open on the official Legistar site]({r['Web Link']})")
        st.markdown(f"**Topic tags:** {r.get('Topic tags','') or '—'}  \n**Boroughs named:** {r.get('Boroughs named','')}")

        st.markdown("### 🔬 Full policy analysis")
        _an = st.session_state.get("analyses", {}).get(mid)
        if not anthropic_key.strip():
            st.info("➕ Add your **Anthropic API key** in the ⚙️ controls panel to generate a full analysis "
                    "(what it does, who supports/opposes, political, district/borough/city, fiscal, why it exists, "
                    "what happens if passed).")
        else:
            cg = st.columns([1, 3])
            if cg[0].button("Generate analysis" if not _an else "↻ Regenerate", key=f"an_btn_{mid}"):
                with st.spinner("Analyzing this bill (NYC Open Data + AI)… ~10–20s"):
                    try:
                        r_an = dict(r); r_an["_sponsor_objs"] = sp
                        _an = AIImpact("claude-haiku-4-5-20251001", api_key=anthropic_key.strip()).analyze(r_an, tx, build_data_context(r))
                        st.session_state.setdefault("analyses", {})[mid] = _an
                    except Exception as e:
                        st.error(f"{type(e).__name__}: {e}")
            if not _an:
                cg[1].caption("On-demand, so opening a bill stays fast.")
        if _an:
            st.caption("Grounded in the bill text and any retrieved NYC Open Data. Inference, not official — verify "
                       "figures with OMB / IBO / agency sources.")
            st.markdown(_an)
        if sp:
            st.markdown("**Sponsors (signature order):**")
            st.dataframe(pd.DataFrame([{"#": s.get("MatterSponsorSequence"), "Sponsor": s.get("MatterSponsorName"),
                "Role": "Prime" if s.get("MatterSponsorSequence") == 0 else "Co-sponsor"} for s in sp]),
                hide_index=True, use_container_width=True)
        if hi:
            st.markdown("**Action history (incl. vote results):**")
            st.dataframe(pd.DataFrame([{"Date": _date(h.get("MatterHistoryActionDate")),
                "Action": (h.get("MatterHistoryActionName") or "").strip(),
                "By": h.get("MatterHistoryActionBodyName"), "Result": h.get("MatterHistoryPassedFlagName")}
                for h in sorted(hi, key=lambda x: x.get("MatterHistoryActionDate") or "")]),
                hide_index=True, use_container_width=True)
        with st.expander("🗳️ Roll-call votes — who voted yes / no"):
            st.caption("Pulls the recorded committee / Stated Meeting roll calls for this bill from Legistar. "
                       "Most bills only have votes once they advance out of committee.")
            if st.button("Load roll-call votes", key=f"votes_btn_{mid}"):
                with st.spinner("Fetching votes from Legistar…"):
                    st.session_state.setdefault("votes_cache", {})[mid] = fetch_votes(mid)
            ve = st.session_state.get("votes_cache", {}).get(mid)
            if ve is not None:
                if not ve:
                    st.caption("No roll-call votes are recorded for this bill yet.")
                for ev in ve:
                    head = f"**{ev['date']} · {ev['body'] or 'Council'}** — {ev['action']}"
                    if ev["result"]:
                        head += f"  ·  **{ev['result']}**"
                    st.markdown(head)
                    if ev["tally"]:
                        st.caption("  ·  ".join(f"{k}: {v}" for k, v in ev["tally"].items()))
                    if ev["votes"]:
                        st.dataframe(pd.DataFrame(ev["votes"]), hide_index=True,
                                     use_container_width=True, height=min(420, 60 + 28 * len(ev["votes"])))
        if at:
            st.markdown("**Attachments:**")
            for a in at:
                st.markdown(f"- [{a.get('MatterAttachmentName')}]({a.get('MatterAttachmentHyperlink')})")
        ab = bundle.get("ai_map", {}).get(mid)
        if ab:
            st.markdown("**AI impact read (analysis, not official):**")
            st.write(f"- **Purpose:** {ab.get('purpose','')}")
            st.write(f"- **Who it affects:** {ab.get('affects','')}")
            st.write(f"- **Where it lands:** {ab.get('local', ab.get('d49',''))}")
        if tx:
            with st.expander("Full bill text"):
                st.text(tx)



# ---------------- MEMBERS ----------------
with t_members:
    if not bundle:
        need_data()
    elif not any(r.get("_sponsor_names") for r in rows):
        st.info("Sponsor data isn't loaded. Turn on **Include sponsors** in the ⚙️ controls panel, or use the "
                "**By a Council Member** scope, then reload.")
    else:
        mem = st.text_input("Council Member last name")
        if mem.strip():
            mb = member_bills(rows, mem); prime = member_prime_count(mb, mem)
            c = st.columns(3)
            c[0].metric("Bills (on)", len(mb)); c[1].metric("As prime", prime); c[2].metric("As co-sponsor", len(mb) - prime)
            pc = pillar_counts(mb)
            if pc:
                st.subheader("Their bills by policy topic"); st.bar_chart(pd.Series(pc).sort_values(ascending=False))
            df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_") and k != "Topic Tags"} for r in mb])
            if not df.empty:
                st.dataframe(df, use_container_width=True, height=340,
                    column_config={"Web Link": st.column_config.LinkColumn("Legistar", display_text="Open")})
            last = mem.split()[-1] if mem.split() else mem
            coal = coalition_counts(mb, member=last)
            if coal:
                st.subheader(f"Who {mem} co-sponsors with most"); st.bar_chart(pd.Series(coal))

# ---------------- DOSSIER ----------------
with t_dossier:
    st.subheader("📕 Council Member dossier")
    st.caption("A profile of any member's record for the selected **Year** (controls panel), with optional AI analysis. "
               "Loads on its own.")
    members = get_directory()
    if not members:
        manual = st.text_input("Directory unavailable — type a member's last name", "", key="dossier_manual")
        members = [manual.strip()] if manual.strip() else []
    who = st.selectbox("Council Member", members, key="dossier_member") if members else None
    run_ai = st.checkbox("Include AI analysis (uses the Anthropic key in the controls panel)", value=bool(anthropic_key.strip()))
    _have_sponsors = bool(bundle) and any(r.get("_sponsor_names") for r in rows)
    if _have_sponsors:
        st.caption("⚡ Fast mode: building from the data already loaded (turn on **Include sponsors** when loading a "
                   "year to keep this instant). Otherwise it scans that member's year live.")
    if who and st.button("Build dossier", type="primary"):
        try:
            if _have_sponsors and member_bills(rows, who):
                mb0 = member_bills(rows, who)
                dd = {"member": who, "rows": mb0, "stats": dossier_stats(mb0, who)}
            else:
                with st.spinner(f"Scanning {who}'s {year} record (faster now; cached after)..."):
                    dd = build_member_dossier(who, year)
            st.session_state["dossier"] = dd; st.session_state["dossier_ai"] = ""
            if run_ai and anthropic_key.strip():
                with st.spinner("Writing AI analysis..."):
                    st.session_state["dossier_ai"] = make_dossier_ai(who, dd["stats"], anthropic_key.strip())
        except Exception as e:
            st.error(f"{type(e).__name__}: {e}")
    dd = st.session_state.get("dossier")
    if dd:
        stats = dd["stats"]; mb = dd["rows"]; member = dd["member"]
        st.markdown(f"## {member}")
        # Introductions-only counts (intro.nyc methodology): prime = "introduced", co = "sponsored"
        _intros = [r for r in mb if r.get("Type") == "Introduction"]
        _intro_prime = sum(1 for r in _intros if member.lower() in (r.get("Prime Sponsor", "") or "").lower())
        _intro_co = len(_intros) - _intro_prime
        c = st.columns(4)
        c[0].metric("Introduced (prime)", stats["as_prime"]); c[1].metric("Co-sponsored", stats["as_cosponsor"])
        c[2].metric("Enacted/passed", stats["by_status"]["passed"]); c[3].metric("On — total", stats["bills_on"])
        st.caption(f"**Introductions only** (matches intro.nyc): introduced **{_intro_prime}**, co-sponsored "
                   f"**{_intro_co}**.  The totals above also include resolutions and land-use items.")
        import unicodedata as _ud
        _slug = "-".join(_ud.normalize("NFKD", member).encode("ascii", "ignore").decode()
                         .lower().replace(".", "").replace("'", "").split())
        st.markdown(f"🔗 Cross-check this member on [intro.nyc](https://intro.nyc/councilmembers/{_slug}) "
                    f"(same official data source).")
        _fails = dd.get("scan_failures") or 0
        if _fails:
            st.warning(f"⚠️ {_fails} sponsor lookups failed during the live scan, so this may **undercount**. "
                       f"Reload (or load **All legislation + Include sponsors** for the year) to get a complete, cached set.")
        cc = st.columns(2)
        if stats["by_topic"]:
            cc[0].subheader("Policy topics"); cc[0].bar_chart(pd.Series(stats["by_topic"]).sort_values(ascending=False))
        if stats["top_coalition"]:
            cc[1].subheader("Top coalition partners"); cc[1].bar_chart(pd.Series(stats["top_coalition"]))
        cd = st.columns(2)
        _sc = status_counts(mb)
        if _sc:
            cd[0].subheader("Outcomes"); cd[0].bar_chart(pd.Series(_sc))
        import collections as _co
        _months = _co.Counter((r.get("_intro_raw") or "")[:7] for r in mb if r.get("_intro_raw"))
        if len(_months) > 1:
            cd[1].subheader("Bills introduced over time")
            cd[1].line_chart(pd.Series(dict(sorted(_months.items()))))
        ai_txt = st.session_state.get("dossier_ai", "")
        if ai_txt:
            st.markdown("### 🧠 AI analysis")
            st.caption("Generated by AI from this member's public legislative record only. Inference, not official — "
                       "reflects sponsorship activity (not floor votes) and may be incomplete.")
            st.write(ai_txt)
        elif run_ai and not anthropic_key.strip():
            st.info("Add your Anthropic key in the ⚙️ controls panel to include the AI write-up.")
        st.markdown("### Prime-sponsored bills")
        pm = [r for r in mb if member.lower() in (r.get("Prime Sponsor", "") or "").lower()]
        df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_") and k != "Topic Tags"} for r in pm])
        if not df.empty:
            st.dataframe(df, use_container_width=True, height=320,
                column_config={"Web Link": st.column_config.LinkColumn("Legistar", display_text="Open")})

# ---------------- COMPARE ----------------
with t_compare:
    if not bundle or not any(r.get("_sponsor_names") for r in rows):
        st.info("Turn on **Include sponsors** (or use a member scope) so two members can be compared.")
    else:
        c = st.columns(2)
        a = c[0].text_input("Member A (last name)", ""); b = c[1].text_input("Member B (last name)", "")
        if a.strip() and b.strip():
            ma, mb = member_bills(rows, a), member_bills(rows, b)
            comp = pd.DataFrame({"Member": [a, b], "Bills (on)": [len(ma), len(mb)],
                "As prime": [member_prime_count(ma, a), member_prime_count(mb, b)],
                "Alive": [overview_general(ma)["alive"], overview_general(mb)["alive"]],
                "Passed": [overview_general(ma)["passed"], overview_general(mb)["passed"]],
                "Dead/filed": [overview_general(ma)["dead"], overview_general(mb)["dead"]]})
            st.dataframe(comp, hide_index=True, use_container_width=True)

# ---------------- COALITIONS (co-sponsorship network) ----------------
def coalition_html(nodes, edges):
    import json as _json
    return """
<div id="cstat" style="font:13px Arial;color:#9fb3d1;padding:4px 2px;">Drawing network…</div>
<div id="net" style="height:560px;border-radius:12px;background:#0a0f1c;border:1px solid #1e2a44;"></div>
<script>
function draw(){
  if (typeof vis === 'undefined') { document.getElementById('cstat').innerHTML =
     'Could not load the graph library (network blocked the CDN). The Strongest-partnerships table below still works.'; return; }
  var nodes = new vis.DataSet(%s);
  var edges = new vis.DataSet(%s);
  var net = new vis.Network(document.getElementById('net'), {nodes:nodes, edges:edges}, {
    nodes:{shape:'dot', scaling:{min:6,max:42, label:{min:11,max:22}},
           font:{color:'#e7eefb', size:13, face:'Arial'},
           color:{background:'#2563eb', border:'#7dd3fc', highlight:{background:'#3b82f6', border:'#bae6fd'}}},
    edges:{color:{color:'#33507f', highlight:'#7dd3fc'}, scaling:{min:1,max:9},
           smooth:{type:'continuous'}, selectionWidth:2},
    physics:{stabilization:{iterations:120, fit:true},
             barnesHut:{gravitationalConstant:-9000, springLength:135, springConstant:0.03}},
    interaction:{hover:true, tooltipDelay:120}
  });
  net.once('stabilizationIterationsDone', function(){ net.setOptions({physics:false}); 
    document.getElementById('cstat').innerHTML = 'Drag dots to rearrange · hover to highlight · scroll to zoom.'; });
}
var s=document.createElement('script');
s.src='https://cdn.jsdelivr.net/npm/vis-network@9.1.9/standalone/umd/vis-network.min.js';
s.onload=draw;
s.onerror=function(){ document.getElementById('cstat').innerHTML='Could not load the graph library from the CDN. The table below still works.'; };
document.head.appendChild(s);
</script>""" % (_json.dumps(nodes), _json.dumps(edges))

with t_net:
    if not bundle or not any(r.get("_sponsor_names") for r in rows):
        st.info("Turn on **Include sponsors** in the ⚙️ controls panel (or load a year of all legislation with sponsors) "
                "so co-sponsorship coalitions can be drawn.")
    else:
        st.subheader("🤝 Co-sponsorship coalitions")
        st.caption("Who works with whom across the loaded bills. The grid below is the clearest view: each cell shows how "
                   "many bills two members co-sponsored — darker means more shared bills.")
        topn = st.slider("Members to include (most active first)", 6, 40, 16, key="net_top")
        members, deg, pair = coalition_matrix(rows, top_members=topn)
        if not members:
            st.warning("No sponsor data in the loaded set yet.")
        else:
            # last-name labels, disambiguated if needed
            def _short(n):
                return n.split()[-1] if n.split() else n
            shorts, seen = [], {}
            for m in members:
                lbl = _short(m); seen[lbl] = seen.get(lbl, 0) + 1
            for m in members:
                lbl = _short(m)
                shorts.append(m if seen[lbl] > 1 else lbl)
            M = pd.DataFrame(0, index=members, columns=members)
            for (a, b), w in pair.items():
                if a in M.index and b in M.columns:
                    M.loc[a, b] = w; M.loc[b, a] = w
            Ms = M.copy(); Ms.index = shorts; Ms.columns = shorts
            mx = max(1, int(M.values.max()))

            def _cell(v):
                if not v:
                    return "background-color:#0a0f1c;color:#26344d"
                t = v / mx
                rr = int(15 + t * (59 - 15)); gg = int(26 + t * (130 - 26)); bb = int(50 + t * (246 - 50))
                return f"background-color:rgb({rr},{gg},{bb});color:#eaf1fb"
            styled = Ms.style
            styled = (styled.map(_cell) if hasattr(styled, "map") else styled.applymap(_cell))
            styled = styled.format(lambda v: "" if not v else int(v))
            st.dataframe(styled, use_container_width=True, height=min(640, 80 + 30 * len(members)))
            st.caption(f"Darkest cell = {mx} shared bills. Most-active member here: **{_short(members[0])}** "
                       f"({deg[members[0]]} bills).")

            pairs = sorted(({"a": a, "b": b, "w": w} for (a, b), w in pair.items()
                            if a in members and b in members), key=lambda x: x["w"], reverse=True)[:15]
            if pairs:
                st.markdown("**Strongest partnerships (most bills co-sponsored together):**")
                st.dataframe(pd.DataFrame([{"Member A": p["a"], "Member B": p["b"], "Shared bills": p["w"]}
                                           for p in pairs]), hide_index=True, use_container_width=True)

            st.markdown("**💡 Why do these coalitions form?**")
            if not anthropic_key.strip():
                st.caption("Add your Anthropic key in the ⚙️ controls panel to get an explanation of the drivers "
                           "(committees, borough delegations, party, caucuses, leadership).")
            elif st.button("Explain the patterns", key="why_coal"):
                with st.spinner("Analyzing the coalition structure…"):
                    try:
                        evidence = {
                            "top_partnerships": [{"a": p["a"], "b": p["b"], "shared_bills": p["w"]} for p in pairs],
                            "most_active_members": [{"member": m, "bills_sponsored": deg[m]} for m in members[:topn]],
                            "session_scope": f"{len(rows)} bills loaded for {loaded_year}",
                        }
                        q = ("Explain why these co-sponsorship coalitions appear in the New York City Council. "
                             "Consider shared committee membership, borough delegations, party (most members are Democrats; "
                             "note the Republican minority), ideological caucuses such as the Progressive Caucus, and "
                             "leadership/whip dynamics. Use the partnership data as evidence, name specific members, and "
                             "search the web for current committee assignments or caucus rosters where that sharpens the "
                             "explanation. Note this reflects only the loaded bills, not the full session.")
                        ans = AIImpact("claude-haiku-4-5-20251001", api_key=anthropic_key.strip()).chat_answer(q, evidence, allow_web=True)
                        st.session_state["coal_why"] = ans
                    except Exception as e:
                        st.session_state["coal_why"] = f"Sorry — {type(e).__name__}: {e}"
            if st.session_state.get("coal_why"):
                st.markdown(st.session_state["coal_why"])
                st.caption("Interpretive analysis based on the loaded data and general knowledge of the Council — verify "
                           "specifics against official committee and caucus rosters.")

            with st.expander("🕸️ Show the network graph instead"):
                minw = st.slider("Min. shared bills for a line", 1, 10, 2, key="net_min")
                nodes, edges = coalition_edges(rows, top_members=topn, min_weight=minw)
                if nodes:
                    st.components.v1.html(coalition_html(nodes, edges), height=580)
                    st.caption(f"{len(nodes)} members · {len(edges)} links. Raise the minimum if it looks crowded.")

# ---------------- DISTRICT MAP ----------------
def district_map_html(highlight=None):
    sel = "null" if not highlight else str(int(highlight))
    return r"""
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/leaflet@1.9.4/dist/leaflet.css"/>
<div id="mstat" style="font:13px Arial;color:#9fb3d1;padding:6px 2px;">Loading district boundaries…</div>
<div id="map" style="height:560px;border-radius:12px;border:1px solid #1e2a44;"></div>
<script>
var SEL = __SEL__;
function go(){
  if (typeof L === 'undefined'){ document.getElementById('mstat').innerHTML='Could not load the map library (CDN blocked).'; return; }
  var map = L.map('map',{scrollWheelZoom:true}).setView([40.70,-73.94],10);
  L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png',
    {attribution:'&copy; OpenStreetMap &copy; CARTO', subdomains:'abcd', maxZoom:19}).addTo(map);
  var SOURCES = [
    'https://services5.arcgis.com/GfwWNkhOj9bNBqoJ/arcgis/rest/services/NYC_City_Council_Districts/FeatureServer/0/query?where=1%3D1&outFields=*&outSR=4326&f=geojson',
    'https://services5.arcgis.com/GfwWNkhOj9bNBqoJ/arcgis/rest/services/NYC_City_Council_Districts_Water_Included/FeatureServer/0/query?where=1%3D1&outFields=*&outSR=4326&f=geojson',
    'https://data.cityofnewyork.us/api/geospatial/6dxp-cfic?method=export&format=GeoJSON'
  ];
  function dnum(p){ if(!p) return null; var ks=['CounDist','coun_dist','council_di','COUNDIST','District','district','councildist']; for(var i=0;i<ks.length;i++){ if(p[ks[i]]!=null) return parseInt(p[ks[i]]); } return null; }
  function style(f){ var d=dnum(f.properties); var on=(SEL!==null && d===SEL);
    return {color: on?'#ffffff':'#39507f', weight: on?2.6:0.7, fillColor: on?'#2563eb':'#1e2a44', fillOpacity: on?0.6:0.22}; }
  var selLayer=null;
  function each(f,layer){ var d=dnum(f.properties);
    layer.bindPopup('<b>Council District '+(d||'?')+'</b>');
    if(SEL!==null && d===SEL) selLayer=layer;
    layer.on('mouseover',function(){layer.setStyle({weight:3,fillOpacity:0.55});});
    layer.on('mouseout',function(){layer.setStyle(style(f));}); }
  function attempt(i){
    if(i>=SOURCES.length){ document.getElementById('mstat').innerHTML='Could not load live district boundaries right now (network or dataset URL). Everything else in the app is unaffected — tell me and I can point this at a specific dataset.'; return; }
    fetch(SOURCES[i]).then(function(r){ if(!r.ok) throw 0; return r.json(); }).then(function(g){
      var layer=L.geoJSON(g,{style:style,onEachFeature:each}).addTo(map);
      try{ if(selLayer){ map.fitBounds(selLayer.getBounds(),{padding:[20,20]}); } else { map.fitBounds(layer.getBounds(),{padding:[12,12]}); } }catch(e){}
      document.getElementById('mstat').innerHTML = (SEL!==null)
        ? ('NYC City Council districts — <b style="color:#3b82f6">District '+SEL+'</b> highlighted. Click any district for its number.')
        : 'All 51 NYC City Council districts. Click any district for its number, or pick one above to highlight it.';
    }).catch(function(e){ attempt(i+1); });
  }
  attempt(0);
}
var ls=document.createElement('script');
ls.src='https://cdn.jsdelivr.net/npm/leaflet@1.9.4/dist/leaflet.js';
ls.onload=go; ls.onerror=function(){ document.getElementById('mstat').innerHTML='Could not load the map library from the CDN.'; };
document.head.appendChild(ls);
</script>""".replace("__SEL__", sel)

with t_map:
    st.subheader("🗺️ NYC City Council districts")
    st.caption("All 51 Council districts, citywide. Pick any district to highlight it, and click any district on the "
               "map for its number. Boundaries load live from NYC Open Data in your browser.")
    hl = st.selectbox("Highlight a district (optional)", ["(none)"] + [str(i) for i in range(1, 52)], key="map_hl")
    st.components.v1.html(district_map_html(None if hl == "(none)" else int(hl)), height=620)
    st.caption("Reference map. If the boundaries don't appear, your network may be blocking the data source — let me "
               "know and I'll wire it to a specific dataset URL.")


# ---------------- OVERVIEW ----------------
with t_over:
    if not bundle:
        need_data()
    else:
        s = overview_general(rows)
        a = st.columns(4)
        a[0].metric("Bills loaded", s["total"]); a[1].metric("Alive (in committee)", s["alive"])
        a[2].metric("Passed", s["passed"]); a[3].metric("Dead / filed", s["dead"])
        tc = type_counts(rows)
        if tc:
            st.caption("By type: " + " · ".join(f"{k}: {v}" for k, v in sorted(tc.items(), key=lambda x: -x[1])))
        with open("/tmp/legislation.xlsx", "rb") as fh:
            st.download_button("⬇️ Download everything as Excel", fh.read(), "NYC_Council_legislation.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        c1, c2 = st.columns(2)
        pc = pillar_counts(rows)
        if pc:
            c1.subheader("By policy topic"); c1.bar_chart(pd.Series(pc).sort_values(ascending=False))
        sc = status_counts(rows)
        if sc:
            c2.subheader("By status"); c2.bar_chart(pd.Series(sc))

# ---------------- WHAT CHANGED ----------------
with t_changes:
    if not bundle:
        need_data()
    else:
        st.subheader("Changes since the last load (while this app stays awake)")
        ch = bundle.get("changes", [])
        if ch:
            st.dataframe(pd.DataFrame(ch, columns=["File", "Change", "Field", "Was", "Now", "When (last modified UTC)"]),
                         hide_index=True, use_container_width=True)
            st.caption("**SIGNED ON +** = new co-sponsor, **AMENDED (text)** = the bill's language/version changed "
                       "(* → A → B), **STATUS** = moved in the process. 'When' is Legistar's last-modified timestamp.")
        else:
            st.info("No diff yet — this compares against the previous load while the app stays awake. The list below "
                    "always works.")
        st.divider()
        st.subheader("🕐 Recently changed (by Legistar last-modified time)")
        recent = sorted([r for r in rows if r.get("Last Modified (UTC)")],
                        key=lambda r: r["Last Modified (UTC)"], reverse=True)[:60]
        if recent:
            rdf = pd.DataFrame([{"File": r["File"], "Type": r["Type"], "Status": r["Status"],
                                 "Version": r.get("Version", "*"),
                                 "Last modified (UTC)": (r["Last Modified (UTC)"] or "")[:16].replace("T", " "),
                                 "Web Link": r["Web Link"]} for r in recent])
            st.dataframe(rdf, use_container_width=True, height=420,
                column_config={"Web Link": st.column_config.LinkColumn("Legistar", display_text="Open")})
            st.caption("**Version** shows the bill text: * = original, A/B = amended (the language changed). "
                       "Sort by clicking the 'Last modified' column header.")
        st.caption("For permanent day-to-day tracking that survives restarts, the scheduled version (handed to "
                   "Council IT) keeps a lasting history.")

# ---------------- ABOUT ----------------
with t_about:
    st.subheader("About this tool")
    st.markdown("""
**NYC Legislative Intelligence** tracks the legislation, people, elections, and votes that shape New York City —
across **three levels of government** — and turns them into desk-ready briefings.

**📜 Legislation (City Hall)** — live from the NYC Council's official **Legistar** system:
- **Legislation list** — every bill for the chosen year, all types, searchable by number or word.
- **Hearings** — committee schedule, locations, agendas, and outcomes.
- **Bill detail** — sponsors, committee, status, history, roll-call votes, attachments, full text.
- **What changed** — new co-sponsors, amendments, and status moves since the last load.

**🌐 All Levels (Albany + Washington)**
- **State & Federal** — search **NY State** bills (NY Senate Open Legislation API) and track what **NYC's
  U.S. House delegation and senators** are sponsoring in Congress (Congress.gov API).
- **Who governs NYC** — a unified directory of officials at the city, state, and federal levels.
- **Elections & terms** — a deterministic calendar of every office and when it's next on the ballot.

**👥 People & Coalitions** — member records, dossiers, head-to-head compare, co-sponsorship coalitions, district map.

**📰 Briefings & Ideas**
- **Briefing Studio** — *“Bulletpoints for Bureaucrats.”* Turn any bill, member, or topic into a tight,
  plain-English briefing (staff, press-ready, constituent, or one-pager), copy-ready and exportable.
- **Policy Lab** — brainstorm new laws and policies: structured, staff-ready concepts with mechanism,
  sponsor/opposition, fiscal & legal flags, precedents, and a draft intro summary.

**Data vs. analysis.** Facts (bills, sponsors, votes, members) come from official APIs. Anything AI-written is
labeled as **analysis/inference**, is grounded in the provided data, and never invents figures — it names the
source to check. Add your **Anthropic key** in the ⚙️ controls panel to switch on the AI features; the NY State
and Congress bill searches use their own free API keys, entered on the **State & Federal** tab.

_Keys are used only for your live session and are not stored by this app._
""")


# ============================================================================
# v2 — shared helpers for the new tabs
# ============================================================================
LEVEL_BADGE = {"NYC": "b-nyc", "NY State": "b-state", "Federal": "b-fed",
               "U.S. Congress": "b-fed"}


def _badge(text, cls="b-muted"):
    return f'<span class="badge {cls}">{text}</span>'


def _level_badge(level):
    return _badge(level, LEVEL_BADGE.get(level, "b-muted"))


def _get_llm(smart=False):
    """Build the shared LLM client from the key in the controls panel."""
    key = (anthropic_key or "").strip() or os.environ.get("ANTHROPIC_API_KEY")
    model = _llm.SMART_MODEL if smart else _llm.FAST_MODEL
    return _llm.LLM(api_key=key, model=model)


@st.cache_data(ttl=86400, show_spinner=False)
def load_federal_delegation():
    """NYC's U.S. House members + both NY senators (no key needed)."""
    if not _cong:
        return []
    legs = _cong.load_legislators()
    return [_people.federal_profile(p) for p in _cong.nyc_delegation(legs)]


@st.cache_data(ttl=1800, show_spinner=False)
def nys_search(term, yr, key):
    if not _nys:
        return []
    return _nys.NYStateClient(api_key=key).search_bills(term, year=yr, limit=50)


@st.cache_data(ttl=1800, show_spinner=False)
def nys_members(yr, key, chamber):
    if not _nys:
        return []
    return _nys.NYStateClient(api_key=key).members(year=yr, chamber=chamber)


@st.cache_data(ttl=1800, show_spinner=False)
def congress_member_bills(bioguide, kind, key):
    if not _cong:
        return []
    return _cong.CongressClient(api_key=key).member_legislation(bioguide, kind=kind)


def _brief_download_row(md, title, key_prefix):
    """Copy box + Markdown / print-HTML / Excel download buttons for a briefing."""
    st.code(md, language="markdown")
    cols = st.columns(3)
    cols[0].download_button("⬇️ Markdown", md, f"{title}.md", "text/markdown",
                            key=f"{key_prefix}_md", use_container_width=True)
    html = _brief.print_html(_brief.md_to_html(md), title=title)
    cols[1].download_button("🖨️ Print / PDF (HTML)", html, f"{title}.html", "text/html",
                            key=f"{key_prefix}_html", use_container_width=True)
    try:
        from openpyxl import Workbook as _WB
        wb = _WB(); ws = wb.active; ws.title = "Briefing"
        ws.append(["Section", "Content"])
        for sec, txt in _brief.briefing_to_rows(md):
            ws.append([sec, txt])
        ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 90
        path = "/tmp/briefing.xlsx"; wb.save(path)
        with open(path, "rb") as fh:
            cols[2].download_button("⬇️ Excel", fh.read(), f"{title}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{key_prefix}_xlsx", use_container_width=True)
    except Exception:
        cols[2].caption("Excel export unavailable.")
    if st.button("💾 Save to Knowledge base", key=f"{key_prefix}_save"):
        _mem().save_item("briefing", title, md)
        st.success("Saved — find it in 🧠 Knowledge & Memory.")


# ============================================================================
# 🏛️ COMMAND CENTER
# ============================================================================
with t_home:
    st.markdown(
        '<div class="hero"><h1>Command Center</h1>'
        '<p>One desk for every level of government that touches New York City — '
        'and the briefings that make it repeatable.</p></div>', unsafe_allow_html=True)

    if bundle and rows:
        s = overview_general(rows)
        m = st.columns(4)
        m[0].metric("Bills loaded", s["total"])
        m[1].metric("Alive (in committee)", s["alive"])
        m[2].metric("Passed / enacted", s["passed"])
        recent_n = len([r for r in rows if r.get("Last Modified (UTC)")])
        m[3].metric("With activity", recent_n)
        st.caption(f"Loaded scope: **{loaded_year or '—'}** · "
                   f"{'sponsors included' if any(r.get('_sponsor_names') for r in rows) else 'sponsors not loaded'}. "
                   "Use **Briefing Studio** to turn any of these into a one-pager.")
        cc = st.columns(2)
        pc = pillar_counts(rows)
        if pc:
            cc[0].markdown("**Top policy topics loaded**")
            cc[0].bar_chart(pd.Series(dict(sorted(pc.items(), key=lambda x: -x[1])[:8])))
        recent = sorted([r for r in rows if r.get("Last Modified (UTC)")],
                        key=lambda r: r["Last Modified (UTC)"], reverse=True)[:6]
        if recent:
            cc[1].markdown("**Most recently active bills**")
            cc[1].dataframe(pd.DataFrame([{"File": r["File"], "Status": r["Status"],
                "Updated": (r["Last Modified (UTC)"] or "")[:10]} for r in recent]),
                hide_index=True, use_container_width=True)
    else:
        st.info("**Start here:** open the ⚙️ Data controls panel above, pick **All legislation + 2026**, and press "
                "**Load data**. Then every tab — and the Briefing Studio — comes alive. The **All Levels**, "
                "**Elections**, and **Policy Lab** tabs work even before you load City Council data.")

    st.divider()
    st.markdown('<span class="kicker">Three levels of government</span>', unsafe_allow_html=True)
    bs = _people.branch_summary()
    lc = st.columns(3)
    for i, lvl in enumerate(_people.LEVELS):
        branches = bs.get(lvl, {})
        body = "".join(f'<div class="chip">{b}: {n}</div>' for b, n in branches.items())
        lc[i].markdown(
            f'<div class="card {"pcard" if lvl=="NYC" else "pcard state" if lvl=="NY State" else "pcard fed"}">'
            f'<h4>{_level_badge(lvl)} {lvl}</h4><div style="margin-top:8px">{body}</div></div>',
            unsafe_allow_html=True)

    st.markdown('<span class="kicker">On the ballot soon</span>', unsafe_allow_html=True)
    cal = _people.election_calendar(_dt.date.today().year, years=4)
    if cal:
        st.dataframe(pd.DataFrame([{"Year": r["Year"], "Level": r["Level"], "Office": r["Office"],
                                    "In (yrs)": r["In"]} for r in cal[:12]]),
                     hide_index=True, use_container_width=True, height=300)
    st.caption("Full calendar and who's up for re-election is on the **Elections & terms** tab.")


# ============================================================================
# 🏙️ STATE & FEDERAL
# ============================================================================
with t_gov:
    st.subheader("🏙️ State & Federal action that affects NYC")
    st.caption("Albany and Washington set the rules NYC lives under — from housing and transit funding to the "
               "criminal code. Track NY State bills and what NYC's congressional delegation is doing.")
    kc = st.columns(2)
    nys_key = kc[0].text_input("NY State Open Legislation API key",
        st.session_state.get("nys_key", ""), type="password", key="nys_key_in",
        help="Free at legislation.nysenate.gov/public/subscribe")
    cong_key = kc[1].text_input("Congress.gov API key",
        st.session_state.get("cong_key", ""), type="password", key="cong_key_in",
        help="Free at api.congress.gov/sign-up")
    st.session_state["nys_key"] = nys_key
    st.session_state["cong_key"] = cong_key

    g_state, g_fed = st.tabs(["🟣 NY State legislation", "🔴 U.S. Congress (NYC delegation)"])

    with g_state:
        st.markdown(f"{_level_badge('NY State')} &nbsp; Search NY State bills for the "
                    f"**{_nys.session_year(int(year) if str(year).isdigit() else 2025) if _nys else '2025'}** session.",
                    unsafe_allow_html=True)
        sc = st.columns([4, 1])
        nq = sc[0].text_input("Search NY State bills (e.g. 'housing NYC', 'MTA', 'rent')", key="nys_q")
        sc[1].write(""); sc[1].write("")
        go_nys = sc[1].button("Search", key="nys_go", use_container_width=True)
        if go_nys:
            if not nys_key.strip():
                st.warning("Add your **NY State Open Legislation API key** above to search Albany. It's free.")
            elif not nq.strip():
                st.warning("Type something to search for.")
            else:
                yr = int(year) if str(year).isdigit() else 2025
                try:
                    with st.spinner("Searching Albany…"):
                        st.session_state["nys_results"] = nys_search(nq.strip(), yr, nys_key.strip())
                except Exception as e:
                    st.error(f"NY State search failed: {type(e).__name__}: {e}")
        res = st.session_state.get("nys_results")
        if res:
            st.caption(f"{len(res)} NY State bills")
            st.dataframe(pd.DataFrame([{"Bill": r["File"], "Type": r["Type"], "Title": r["Title"][:120],
                "Sponsor": r["Sponsor"], "Status": r["Status"], "Web Link": r["Web Link"]} for r in res]),
                use_container_width=True, height=440,
                column_config={"Web Link": st.column_config.LinkColumn("Open", display_text="nysenate.gov")})
            st.session_state["gov_briefable"] = res
            st.caption("Send any of these to **Briefing Studio → State/Federal bill** for a one-pager.")
        elif res == []:
            st.info("No results (or the search couldn't reach Albany). Try broader terms.")

    with g_fed:
        st.markdown(f"{_level_badge('Federal')} &nbsp; NYC's representation in Washington — the House members whose "
                    "districts sit in the five boroughs, plus both NY senators.", unsafe_allow_html=True)
        deleg = load_federal_delegation()
        if not deleg:
            st.info("Couldn't load the delegation roster right now (the public dataset was unreachable). "
                    "It loads automatically in a normal deployment; try again shortly.")
        else:
            names = {f"{d['name']} — {d['seat']} ({d['party'][:1]})": d for d in deleg}
            pick = st.selectbox("Pick a member of the delegation", list(names.keys()), key="fed_pick")
            d = names[pick]
            info = st.columns(3)
            info[0].metric("Chamber", d["chamber"].replace("U.S. ", ""))
            info[1].metric("Party", d["party"] or "—")
            info[2].metric("In office since", d["extra"].get("since", "—"))
            if not cong_key.strip():
                st.caption("Add your **Congress.gov API key** above to pull this member's sponsored & "
                           "cosponsored bills in the current Congress.")
            else:
                kind = st.radio("Show", ["sponsored", "cosponsored"], horizontal=True, key="fed_kind")
                bio = d["extra"].get("bioguide", "")
                if bio and st.button("Load their bills", key="fed_bills_go"):
                    try:
                        with st.spinner("Pulling from Congress.gov…"):
                            st.session_state["fed_bills"] = congress_member_bills(bio, kind, cong_key.strip())
                    except Exception as e:
                        st.error(f"Congress.gov failed: {type(e).__name__}: {e}")
                fb = st.session_state.get("fed_bills")
                if fb:
                    st.dataframe(pd.DataFrame([{"Bill": b["File"], "Title": (b["Title"] or "")[:120],
                        "Latest action": b["Status"][:80], "Date": b.get("Action date", ""),
                        "Web Link": b["Web Link"]} for b in fb]),
                        use_container_width=True, height=380,
                        column_config={"Web Link": st.column_config.LinkColumn("Open", display_text="congress.gov")})
                    st.session_state["gov_briefable"] = fb
                elif fb == []:
                    st.info("No bills returned for that selection yet.")


# ============================================================================
# 👤 WHO GOVERNS NYC (unified directory)
# ============================================================================
with t_dir:
    st.subheader("👤 Who governs NYC — unified directory")
    st.caption("Every level, one place. Officials' names come live from official sources, so the roster stays current "
               "through elections.")
    lvl = st.radio("Level", ["Federal", "NY State", "NYC Council"], horizontal=True, key="dir_level")

    if lvl == "Federal":
        deleg = load_federal_delegation()
        if not deleg:
            st.info("The public congressional dataset was unreachable just now; it loads automatically in a normal "
                    "deployment.")
        else:
            st.caption(f"{len(deleg)} members — NYC's House delegation + both NY senators.")
            for d in deleg:
                ex = d["extra"]
                extra_bits = []
                if ex.get("since"): extra_bits.append(f"since {ex['since']}")
                if d.get("term_end"): extra_bits.append(f"term ends {d['term_end']}")
                if ex.get("twitter"): extra_bits.append(f"@{ex['twitter']}")
                meta = " · ".join([d["party"]] + extra_bits)
                link = f' · <a href="{d["source"]}">official site</a>' if d.get("source") else ""
                st.markdown(
                    f'<div class="card pcard fed"><h4>{d["name"]} '
                    f'<span class="badge b-fed">{d["seat"]}</span></h4>'
                    f'<div class="meta">{d["chamber"]} · {meta}{link}<br>{d.get("contact","")}</div></div>',
                    unsafe_allow_html=True)

    elif lvl == "NY State":
        nkey = st.session_state.get("nys_key", "")
        if not nkey.strip():
            st.info("Add your **NY State Open Legislation API key** on the **State & Federal** tab to load Albany "
                    "members. (It's free.) Meanwhile: NYC sends members to both the **State Senate** and the "
                    "**State Assembly** — search bills by their names on that tab.")
        else:
            ch = st.radio("Chamber", ["State Senate", "State Assembly"], horizontal=True, key="dir_nys_ch")
            yr = int(year) if str(year).isdigit() else 2025
            try:
                with st.spinner("Loading Albany roster…"):
                    mem = nys_members(yr, nkey.strip(),
                                      "SENATE" if ch == "State Senate" else "ASSEMBLY")
            except Exception as e:
                mem = []; st.error(f"{type(e).__name__}: {e}")
            if mem:
                st.caption(f"{len(mem)} members in the {ch} (statewide — NYC districts included).")
                st.dataframe(pd.DataFrame([{"Name": m["name"], "District": m["district"],
                    "Party": _people._party(m["party"])} for m in
                    sorted(mem, key=lambda x: (x["district"] or 0))]),
                    hide_index=True, use_container_width=True, height=440)
            else:
                st.info("No members returned (or Albany was unreachable).")

    else:  # NYC Council
        members = get_directory()
        if not members:
            st.info("The Council roster loads from Legistar; it was unreachable just now. Load any City Council data "
                    "from the ⚙️ panel and it will populate.")
        else:
            st.caption(f"{len(members)} current Council Members. Open the **People & Coalitions → Dossier** tab for any "
                       "member's full legislative record.")
            cols = st.columns(3)
            for i, nm in enumerate(members):
                cols[i % 3].markdown(
                    f'<div class="card pcard"><h4>{nm}</h4>'
                    f'<div class="meta">{_level_badge("NYC")} NYC City Council</div></div>',
                    unsafe_allow_html=True)


# ============================================================================
# 🗳️ ELECTIONS & TERMS
# ============================================================================
with t_elect:
    st.subheader("🗳️ Elections & terms — what's on the ballot, and when")
    st.caption("A deterministic calendar built from each office's fixed election cycle (no guessing about "
               "outcomes). Cross-check specific dates and any special elections with the NYC/NYS Boards of Elections.")
    this_year = _dt.date.today().year
    yy = st.number_input("Show elections starting from year", min_value=2024, max_value=2040,
                         value=this_year, step=1, key="elect_year")
    up = _people.offices_up_this_year(int(yy))
    if up:
        st.markdown(f"**On the ballot in {int(yy)}:**")
        st.markdown(" ".join(_badge(o["office"], LEVEL_BADGE.get(o["level"], "b-muted"))
                             for o in up), unsafe_allow_html=True)
    else:
        st.caption(f"No regularly-scheduled covered offices open in {int(yy)} (special elections aside).")
    st.divider()
    cal = _people.election_calendar(int(yy), years=8)
    lvlf = st.multiselect("Filter by level", _people.LEVELS, key="elect_lvl")
    show = [r for r in cal if not lvlf or r["Level"] in lvlf]
    st.dataframe(pd.DataFrame([{"Year": r["Year"], "In (yrs)": r["In"], "Level": r["Level"],
        "Office": r["Office"], "Branch": r["Branch"], "Term (yrs)": r["Term (yrs)"]} for r in show]),
        hide_index=True, use_container_width=True, height=420)

    st.divider()
    st.markdown("**Federal delegation — who's up next**")
    deleg = load_federal_delegation()
    if deleg:
        st.dataframe(pd.DataFrame([{"Member": d["name"], "Seat": d["seat"], "Party": d["party"],
            "Term ends": d["term_end"], "Next election": (d["term_end"][:4] if d["term_end"] else "—")}
            for d in deleg]), hide_index=True, use_container_width=True, height=340)
        st.caption("House members run every 2 years; senators' term-end year is their next race.")
    else:
        st.caption("Delegation roster unavailable right now (loads automatically in a normal deployment).")


# ============================================================================
# 📰 BRIEFING STUDIO — "Bulletpoints for Bureaucrats"
# ============================================================================
with t_brief:
    st.subheader("📰 Briefing Studio — “Bulletpoints for Bureaucrats”")
    st.caption("Turn any bill, member, or topic into a tight, plain-English briefing you can hand to the Council "
               "Member or a reporter. Works as a fact sheet without a key; add your Anthropic key for the polished "
               "write-up and press lines.")
    _llm_client = _get_llm(smart=True)
    if not _llm_client.ready:
        st.info("💡 Add your **Anthropic key** in the ⚙️ controls panel for AI-written briefings. Until then you can "
                "still generate the **fact skeleton** for any loaded bill.")
    audience = st.selectbox("Audience & tone", list(_brief.AUDIENCES.keys()), key="brief_aud")
    mode = st.radio("Briefing subject", ["NYC bill", "State/Federal bill", "Council Member", "Policy topic"],
                    horizontal=True, key="brief_mode")

    if mode == "NYC bill":
        if not bundle:
            need_data()
        else:
            pick = st.selectbox("Pick a bill", [r["File"] for r in rows], key="brief_bill")
            r = next(x for x in rows if x["File"] == pick)
            if st.button("Generate briefing", type="primary", key="brief_go1"):
                with st.spinner("Building the briefing…"):
                    mid = r["MatterId"]
                    det = {}
                    try:
                        sp = r.get("_sponsor_objs") or []
                        tx = bundle.get("text_map", {}).get(mid, "")
                        hi = bundle.get("histories_map", {}).get(mid, [])
                        if not sp and not tx:
                            d0 = fetch_detail(mid)
                            sp = current_sponsors({"MatterVersion": None}, d0.get("sponsors", []))
                            tx = d0.get("text", ""); hi = d0.get("histories", [])
                        det["sponsors"] = [x.get("MatterSponsorName") for x in sp]
                        det["history"] = [f"{_date(h.get('MatterHistoryActionDate'))}: "
                                          f"{(h.get('MatterHistoryActionName') or '').strip()}"
                                          for h in (hi or [])[-6:]]
                        rr = dict(r); rr["_sponsor_objs"] = sp
                        _kctx = _memory.context_for_briefing(_mem(), "bill", r["File"],
                                    topic=(r.get("Topic tags") or "").split(";")[0].strip() or None)
                        _mem().log("brief", "bill", r["File"])
                        md = _brief.bill_briefing(_llm_client, rr, text=tx,
                                                  data_ctx=(build_data_context(r) + "\n\n" + _kctx).strip(),
                                                  audience=audience, detail=det)
                    except Exception as e:
                        md = _brief.template_bill_briefing(r) + f"\n\n> _(detail step failed: {e})_"
                    st.session_state["brief_out"] = (md, pick.replace(" ", "_"))
            if st.session_state.get("brief_out"):
                md, title = st.session_state["brief_out"]
                st.markdown(f'<div class="brief">{_brief.md_to_html(md)}</div>', unsafe_allow_html=True)
                st.divider(); _brief_download_row(md, title, "brief1")

    elif mode == "State/Federal bill":
        pool = st.session_state.get("gov_briefable") or []
        if not pool:
            st.info("Load some bills first on the **State & Federal** tab (search Albany or pull a member's Congress "
                    "bills), then come back here — they'll appear in this dropdown.")
        else:
            labels = {f'{b["File"]} — {(b["Title"] or "")[:70]}': b for b in pool}
            pick = st.selectbox("Pick a bill", list(labels.keys()), key="brief_gov")
            b = labels[pick]
            if st.button("Generate briefing", type="primary", key="brief_go2"):
                with st.spinner("Building the briefing…"):
                    md = _brief.bill_briefing(_llm_client, b, text=b.get("Summary", ""), audience=audience)
                    st.session_state["brief_out2"] = (md, b["File"].replace(" ", "_"))
            if st.session_state.get("brief_out2"):
                md, title = st.session_state["brief_out2"]
                st.markdown(f'<div class="brief">{_brief.md_to_html(md)}</div>', unsafe_allow_html=True)
                st.divider(); _brief_download_row(md, title, "brief2")

    elif mode == "Council Member":
        members = get_directory()
        who = st.selectbox("Council Member", members, key="brief_member") if members else \
            st.text_input("Member last name", key="brief_member_txt")
        if who and st.button("Generate briefing", type="primary", key="brief_go3"):
            with st.spinner("Assembling the member's record…"):
                try:
                    if bundle and member_bills(rows, who):
                        mb = member_bills(rows, who); stats = dossier_stats(mb, who)
                    else:
                        dd = build_member_dossier(who, year); stats = dd["stats"]
                    _notes = [n["note"] for n in _mem().notes_for("member", who)]
                    if _notes:
                        stats = dict(stats); stats["staff_notes"] = _notes
                    _mem().log("brief", "member", who)
                    md = _brief.member_briefing(_llm_client, who, stats, audience=audience)
                except Exception as e:
                    md = f"## {who}\n\n> _(couldn't assemble record: {e})_"
                st.session_state["brief_out3"] = (md, who.replace(" ", "_"))
        if st.session_state.get("brief_out3"):
            md, title = st.session_state["brief_out3"]
            st.markdown(f'<div class="brief">{_brief.md_to_html(md)}</div>', unsafe_allow_html=True)
            st.divider(); _brief_download_row(md, title, "brief3")

    else:  # Policy topic
        if not bundle:
            need_data()
        else:
            topics = sorted({p for r in rows for p in (r.get("Topic tags") or "").split("; ") if p})
            topic = st.selectbox("Policy topic", topics, key="brief_topic") if topics else \
                st.text_input("Topic", key="brief_topic_txt")
            if topic and st.button("Generate briefing", type="primary", key="brief_go4"):
                with st.spinner("Scanning the loaded bills…"):
                    match = [r for r in rows if topic.lower() in (r.get("Topic tags") or "").lower()]
                    ev = {"topic": topic, "matching_count": len(match),
                          "sample": [{"File": r["File"], "Title": (r["Title"] or "")[:90],
                                      "Status": r["Status"], "Prime": r.get("Prime Sponsor", "")}
                                     for r in match[:30]]}
                    md = _brief.topic_briefing(_llm_client, topic, ev, audience=audience)
                    st.session_state["brief_out4"] = (md, f"topic_{topic}".replace(" ", "_"))
            if st.session_state.get("brief_out4"):
                md, title = st.session_state["brief_out4"]
                st.markdown(f'<div class="brief">{_brief.md_to_html(md)}</div>', unsafe_allow_html=True)
                st.divider(); _brief_download_row(md, title, "brief4")


# ============================================================================
# 💡 POLICY LAB — brainstorm new legislation
# ============================================================================
with t_lab:
    st.subheader("💡 Policy Lab — brainstorm new laws & policies")
    st.caption("Describe a problem or a goal. The Lab returns concrete legislative concepts — mechanism, lead agency, "
               "who benefits, who pushes back, fiscal & legal flags, a precedent, and a draft intro summary — the raw "
               "material for a real Introduction.")
    lab_llm = _get_llm(smart=True)
    if not lab_llm.ready:
        st.info("💡 Add your **Anthropic key** in the ⚙️ controls panel to generate ideas.")
    goal = st.text_area("What problem or goal do you want legislation for?",
                        placeholder="e.g. Reduce e-bike battery fires in apartment buildings; expand childcare in the Rockaways",
                        key="lab_goal", height=90)
    lc = st.columns([1, 1, 1])
    ctx = lc[0].text_input("Context (district, angle) — optional", key="lab_ctx")
    n = lc[1].slider("How many ideas", 3, 8, 5, key="lab_n")
    lc[2].write(""); lc[2].write("")
    go = lc[2].button("Brainstorm", type="primary", key="lab_go", use_container_width=True)
    if go:
        if not lab_llm.ready:
            st.warning("Add your Anthropic key in the ⚙️ controls panel first.")
        elif not goal.strip():
            st.warning("Describe the problem or goal first.")
        else:
            with st.spinner("Generating ideas…"):
                try:
                    st.session_state["lab_ideas"] = _lab.ideate(lab_llm, goal.strip(), ctx.strip(), n=n)
                    st.session_state["lab_memo"] = {}
                except Exception as e:
                    st.error(f"{type(e).__name__}: {e}")
    ideas = st.session_state.get("lab_ideas") or []
    if ideas:
        st.caption(f"{len(ideas)} concepts — expand any one for a full memo.")
        for i, idea in enumerate(ideas):
            stars = "★" * int(idea.get("boldness", 3)) + "☆" * (5 - int(idea.get("boldness", 3)))
            st.markdown(
                f'<div class="card"><h4>{idea.get("title","Idea")} '
                f'<span class="badge b-muted">{idea.get("instrument","")}</span> '
                f'<span class="stars">{stars}</span></h4>'
                f'<div class="meta">{idea.get("one_liner","")}</div>'
                f'<div style="margin-top:8px">'
                f'<div class="chip">🏛️ {idea.get("lead_agency","—")}</div>'
                f'<div class="chip">✅ {idea.get("who_benefits","—")[:60]}</div>'
                f'<div class="chip">⚠️ {idea.get("who_pushes_back","—")[:60]}</div></div></div>',
                unsafe_allow_html=True)
            with st.expander(f"Open concept memo — {idea.get('title','Idea')}"):
                if st.button("Write the one-page memo", key=f"lab_refine_{i}"):
                    with st.spinner("Developing the concept…"):
                        st.session_state.setdefault("lab_memo", {})[i] = _lab.refine(
                            lab_llm, idea, context=ctx.strip())
                memo = st.session_state.get("lab_memo", {}).get(i)
                if memo:
                    st.markdown(f'<div class="brief">{_brief.md_to_html(memo)}</div>', unsafe_allow_html=True)
                    st.download_button("⬇️ Markdown", memo,
                        f"{idea.get('title','concept').replace(' ','_')}.md", "text/markdown",
                        key=f"lab_memo_dl_{i}")
                else:
                    st.markdown(f"**Mechanism:** {idea.get('mechanism','—')}")
                    st.markdown(f"**Fiscal flag:** {idea.get('fiscal_flag','—')}")
                    st.markdown(f"**Legal flag:** {idea.get('legal_flag') or '—'}")
                    st.markdown(f"**Precedent:** {idea.get('precedent','—')}")
                    st.markdown(f"**PR angle:** _{idea.get('pr_angle','—')}_")
                    st.markdown(f"**First step:** {idea.get('first_step','—')}")
        try:
            from openpyxl import Workbook as _WB2
            wb = _WB2(); ws = wb.active; ws.title = "Ideas"
            hdrs = list(_lab.ideas_to_rows(ideas)[0].keys()) if ideas else []
            ws.append(hdrs)
            for row_ in _lab.ideas_to_rows(ideas):
                ws.append([row_[h] for h in hdrs])
            wb.save("/tmp/policy_ideas.xlsx")
            with open("/tmp/policy_ideas.xlsx", "rb") as fh:
                st.download_button("⬇️ Download all ideas (Excel)", fh.read(), "policy_ideas.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="lab_xlsx")
        except Exception:
            pass


# ============================================================================
# v2.1 — cache helpers for the four new pillars
# ============================================================================
try:
    from sources import districts as _dist
except Exception:
    _dist = None


@st.cache_data(ttl=86400, show_spinner=False)
def geocode_lookup(address):
    if not _dist:
        return {"ok": False, "reason": "address lookup module unavailable"}
    return _dist.lookup(address)


@st.cache_data(ttl=86400, show_spinner=False)
def federal_committees():
    if not _cong:
        return {}
    return _cong.load_committee_membership()


@st.cache_data(ttl=1800, show_spinner=False)
def nys_bill_full(print_no, yr, key):
    if not _nys:
        return None
    return _nys.NYStateClient(api_key=key).bill(print_no, year=yr)


def _render_profile(profiles_mod, p_llm, level, name, facts, raw):
    """Shared renderer for a member deep-profile card + facts + AI glance."""
    cls = "pcard" if level == "NYC" else ("pcard state" if level == "NY State" else "pcard fed")
    seat = facts.get("seat") or facts.get("chamber") or ""
    dist = facts.get("district")
    seat_badge = f'<span class="badge b-muted">District {dist}</span>' if dist else \
        (f'<span class="badge b-muted">{seat}</span>' if seat else "")
    st.markdown(
        f'<div class="card {cls}"><h4>{_level_badge(level)} {name} {seat_badge}</h4>'
        f'<div class="meta">{facts.get("chamber","")} · {facts.get("party","") or "—"}'
        + (f' · {facts.get("contact")}' if facts.get("contact") else "") + '</div></div>',
        unsafe_allow_html=True)
    frows = profiles_mod.facts_to_rows(facts)
    if frows:
        # Values mix ints and strings; cast to str so Arrow can serialize the column.
        st.dataframe(pd.DataFrame([(k, str(v)) for k, v in frows], columns=["Field", "Value"]),
                     hide_index=True, use_container_width=True)
    if p_llm.ready:
        g = profiles_mod.glance(p_llm, level, name, facts)
        if g:
            st.markdown("**🧠 Record at a glance**")
            st.markdown(f'<div class="brief">{_brief.md_to_html(g)}</div>', unsafe_allow_html=True)
            st.caption("AI summary from the public data shown above — analysis, not an official statement.")
    else:
        st.caption("Add your Anthropic key (⚙️ panel) for an AI 'record at a glance'.")
    src = raw.get("source") if isinstance(raw, dict) else None
    if src:
        st.markdown(f"[Official page]({src})")


@st.cache_data(ttl=3600, show_spinner=False)
def crime_snapshot_cached(since, until, cats, dataset, token):
    if not _od:
        return []
    return _od.crime_snapshot(since, until, categories=list(cats) or None,
                              dataset=dataset, token=(token or None))


def _grounded_figures_panel(key_prefix):
    """Reusable expander: pull sourced NYC crime figures to cite in messaging."""
    with st.expander("📊 Grounded figures — pull live NYC crime data to cite"):
        if not _od:
            st.caption("Live data module unavailable."); return
        st.caption("Report/complaint counts from **NYC Open Data (NYPD Complaint Data)**, each returned with its "
                   "source and date window so you can cite it honestly. These are *reported complaints* (not "
                   "convictions) and recent periods are provisional. Copy the lines you want into the facts box above.")
        gc = st.columns(3)
        today = _dt.date.today()
        since = gc[0].date_input("From", _dt.date(today.year, 1, 1), key=f"{key_prefix}_from")
        until = gc[1].date_input("To", today, key=f"{key_prefix}_to")
        dataset = gc[2].selectbox("Dataset", ["historic", "current"],
            format_func=lambda d: "Historic (all years)" if d == "historic" else "Current YTD (freshest)",
            key=f"{key_prefix}_ds")
        cats = st.multiselect("Categories", list(_od.CATEGORY_MATCH.keys()),
                              default=["Rape", "Sex crimes (other)"], key=f"{key_prefix}_cats")
        token = st.text_input("Socrata app token (optional — raises the rate limit)",
                              st.session_state.get("socrata_token", ""), type="password",
                              key=f"{key_prefix}_token", help="Free at data.cityofnewyork.us (Developer Settings)")
        st.session_state["socrata_token"] = token
        if st.button("Fetch figures", key=f"{key_prefix}_fetch"):
            with st.spinner("Querying NYC Open Data…"):
                try:
                    st.session_state[f"{key_prefix}_snap"] = crime_snapshot_cached(
                        str(since), str(until), tuple(cats), dataset, token.strip())
                except Exception as e:
                    st.error(f"{type(e).__name__}: {e}")
        snap = st.session_state.get(f"{key_prefix}_snap")
        if snap:
            for r in snap:
                st.code(r["citation"], language="text")
        elif snap == []:
            st.info("No figures returned (the dataset may be unreachable, or no matches in that window).")


# ============================================================================
# 🏠 FIND MY REPS — address → every official who represents it
# ============================================================================
with t_reps:
    st.subheader("🏠 Find my representatives")
    st.caption("Type any NYC address. We geocode it (NYC Planning's GeoSearch) and find the City Council, State "
               "Senate, State Assembly, and U.S. House districts that contain it — then match each to the official who "
               "holds that seat. Great for constituent letters and casework.")
    ac = st.columns([4, 1])
    addr = ac[0].text_input("NYC address", placeholder="e.g. 250 Broadway, Manhattan", key="reps_addr")
    ac[1].write(""); ac[1].write("")
    go_reps = ac[1].button("Find reps", type="primary", key="reps_go", use_container_width=True)
    if go_reps and addr.strip():
        with st.spinner("Geocoding and locating districts…"):
            st.session_state["reps_result"] = geocode_lookup(addr.strip())
    res = st.session_state.get("reps_result")
    if res and not res.get("ok"):
        st.warning(res.get("reason", "Could not resolve that address."))
    elif res and res.get("ok"):
        st.markdown(f"📍 **{res['label']}**")
        fed = load_federal_delegation()
        nys_mem = []
        nkey = st.session_state.get("nys_key", "")
        if nkey.strip():
            yr = int(year) if str(year).isdigit() else 2025
            try:
                nys_mem = nys_members(yr, nkey.strip(), "SENATE") + nys_members(yr, nkey.strip(), "ASSEMBLY")
            except Exception:
                nys_mem = []
        reps = _people.match_reps(res["districts"], fed, nys_mem)
        for r in reps:
            cls = "pcard" if r["level"] == "NYC" else ("pcard state" if r["level"] == "NY State" else "pcard fed")
            dist = f"District {r['district']}" if isinstance(r["district"], int) else (r["district"] or "—")
            who = r["member"] or "<i>see official site</i>"
            link = f' · <a href="{r["link"]}">official page</a>' if r.get("link") else ""
            st.markdown(
                f'<div class="card {cls}"><h4>{_level_badge(r["level"])} {r["seat"]} '
                f'<span class="badge b-muted">{dist}</span></h4>'
                f'<div class="meta">{who}{link}</div></div>', unsafe_allow_html=True)
        if not nkey.strip():
            st.caption("💡 Add your **NY State** key on the *State & Federal* tab to fill in the State Senate/Assembly "
                       "member names automatically. Council and House members resolve without any key.")
        st.caption("District boundaries and geocoding come from official public layers (NYC DCP, U.S. Census "
                   "TIGERweb). Verify edge cases against the NYC/NYS Boards of Elections.")


# ============================================================================
# 🗳️ VOTES & DECISIONS — roll-calls across levels
# ============================================================================
with t_votes:
    st.subheader("🗳️ Votes & decisions")
    st.caption("How it actually went down: recorded roll-call votes at each level. NYC committee & Stated Meeting "
               "votes come from Legistar; NY State floor/committee votes from Open Legislation.")
    v_nyc, v_state, v_fed = st.tabs(
        ["🟦 NYC Council roll-calls", "🟣 NY State votes", "🔴 U.S. House votes"])

    with v_nyc:
        if not bundle:
            need_data()
        else:
            pick = st.selectbox("Pick a bill", [r["File"] for r in rows], key="votes_bill")
            r = next(x for x in rows if x["File"] == pick)
            if st.button("Load roll-call votes", key="votes_load_nyc"):
                with st.spinner("Fetching recorded votes from Legistar…"):
                    st.session_state.setdefault("votes2", {})[r["MatterId"]] = fetch_votes(r["MatterId"])
            ve = st.session_state.get("votes2", {}).get(r["MatterId"])
            if ve is not None:
                if not ve:
                    st.info("No roll-call votes recorded for this bill yet — most bills only get votes once they "
                            "advance out of committee.")
                for ev in ve:
                    head = f"**{ev['date']} · {ev['body'] or 'Council'}** — {ev['action']}"
                    if ev.get("result"):
                        head += f"  ·  **{ev['result']}**"
                    st.markdown(head)
                    if ev.get("tally"):
                        st.caption("  ·  ".join(f"{k}: {v}" for k, v in ev["tally"].items()))
                    if ev.get("votes"):
                        st.dataframe(pd.DataFrame(ev["votes"]), hide_index=True, use_container_width=True,
                                     height=min(420, 60 + 28 * len(ev["votes"])))
            else:
                st.caption("Pick a bill and load its votes. (Tip: enacted bills are the most likely to have roll-calls.)")

    with v_state:
        nkey = st.session_state.get("nys_key", "")
        if not nkey.strip():
            st.info("Add your **NY State Open Legislation API key** on the *State & Federal* tab to pull Albany votes.")
        else:
            pc = st.columns([2, 1])
            pn = pc[0].text_input("NY State bill number (e.g. S1234 or A5678)", key="votes_nys_pn")
            pc[1].write(""); pc[1].write("")
            if pc[1].button("Load votes", key="votes_nys_go", use_container_width=True) and pn.strip():
                yr = int(year) if str(year).isdigit() else 2025
                with st.spinner("Fetching from Albany…"):
                    try:
                        st.session_state["nys_bill_full"] = nys_bill_full(pn.strip().upper(), yr, nkey.strip())
                    except Exception as e:
                        st.error(f"{type(e).__name__}: {e}")
            nb = st.session_state.get("nys_bill_full")
            if nb:
                st.markdown(f"**{nb['File']}** — {nb['Title'][:120]}")
                st.caption(f"Status: {nb['Status']}")
                votes = nb.get("Votes") or []
                if not votes:
                    st.info("No recorded votes on this bill yet.")
                for v in votes:
                    st.markdown(f"**{v['date']} · {v['description']}** ({v['type']})")
                    if v["tally"]:
                        st.caption("  ·  ".join(f"{k}: {n}" for k, n in v["tally"].items()))
                    if v["members"]:
                        st.dataframe(pd.DataFrame(v["members"]), hide_index=True, use_container_width=True,
                                     height=min(420, 60 + 26 * len(v["members"])))

    with v_fed:
        st.caption("How NYC's U.S. House delegation voted on a floor roll-call. Enter the year and roll-call number "
                   "(find them at clerk.house.gov → Votes) — we pull the official House Clerk record and filter to "
                   "NYC's members.")
        fc = st.columns([1, 1, 1])
        fyr = fc[0].number_input("Year", min_value=2015, max_value=2035,
                                 value=_dt.date.today().year, step=1, key="fv_year")
        frn = fc[1].number_input("Roll-call #", min_value=1, max_value=2000, value=1, step=1, key="fv_roll")
        fc[2].write(""); fc[2].write("")
        if fc[2].button("Load House vote", key="fv_go", use_container_width=True):
            st.session_state["fed_vote_tried"] = True
            with st.spinner("Fetching from the House Clerk…"):
                try:
                    st.session_state["fed_vote"] = _housevotes.roll_vote(int(fyr), int(frn)) if _housevotes else None
                except Exception as e:
                    st.session_state["fed_vote"] = None; st.error(f"{type(e).__name__}: {e}")
        fv = st.session_state.get("fed_vote")
        if fv is None and st.session_state.get("fed_vote_tried"):
            st.info("Couldn't find that roll-call (check the year/number, or the Clerk site was unreachable).")
        if fv:
            st.markdown(f"**{fv['bill'] or 'Vote'} — {fv['question']}**  ·  Roll {fv['roll']} "
                        f"({fv['congress']}th Congress)")
            st.markdown(f"Result: **{fv['result']}** · {fv['date']}")
            if fv["totals"]:
                st.caption("House total — " + "  ·  ".join(f"{k}: {n}" for k, n in fv["totals"].items()))
            deleg = load_federal_delegation()
            rows_fd, tally = _housevotes.delegation_positions(fv, deleg)
            if tally:
                st.markdown("**NYC delegation:** " + "  ·  ".join(f"{k}: {n}" for k, n in tally.items()))
            if rows_fd:
                st.dataframe(pd.DataFrame(rows_fd), hide_index=True, use_container_width=True,
                             height=min(480, 60 + 26 * len(rows_fd)))
            else:
                st.caption("None of NYC's House members appear on this roll-call (or the roster was unreachable).")
            if fv.get("source_url"):
                st.markdown(f"[Official House Clerk record]({fv['source_url']})")


# ============================================================================
# 🔔 ACTIVITY (ALL LEVELS) — a watchlist + unified change timeline
# ============================================================================
with t_activity:
    st.subheader("🔔 Activity across all levels")
    st.caption("Track specific bills from any level and see what moved. Add bills to the watchlist, then **Refresh & "
               "diff** to catch status changes, amendments, and new actions since you last checked (this session).")
    if "watchlist" not in st.session_state:
        st.session_state["watchlist"] = _store.load("watchlist", {})  # durable across reruns/restarts
    wl = st.session_state["watchlist"]  # key -> {level, file, last_status, title, link}
    if _store.available():
        st.caption("💾 Watchlist is saved to disk — it survives app restarts within this deployment.")

    with st.expander("➕ Add bills to the watchlist", expanded=not wl):
        st.caption("NYC bills come from your loaded set; State/Federal from your latest searches on the *State & "
                   "Federal* tab.")
        addcols = st.columns(3)
        if bundle and rows:
            nyc_pick = addcols[0].selectbox("NYC bill", ["—"] + [r["File"] for r in rows], key="wl_nyc")
            if addcols[0].button("Add NYC bill", key="wl_add_nyc") and nyc_pick != "—":
                r = next(x for x in rows if x["File"] == nyc_pick)
                wl[f"NYC:{r['File']}"] = {"level": "NYC", "file": r["File"], "last_status": r["Status"],
                                          "title": r["Title"], "link": r["Web Link"], "mid": r["MatterId"]}
                _store.save("watchlist", wl)
                st.success(f"Watching {r['File']}")
        else:
            addcols[0].caption("Load NYC data to add city bills.")
        pool = st.session_state.get("gov_briefable") or []
        if pool:
            gp = addcols[1].selectbox("State/Federal bill", ["—"] + [b["File"] for b in pool], key="wl_gov")
            if addcols[1].button("Add State/Fed bill", key="wl_add_gov") and gp != "—":
                b = next(x for x in pool if x["File"] == gp)
                wl[f'{b.get("level","?")}:{b["File"]}'] = {"level": b.get("level", "?"), "file": b["File"],
                    "last_status": b.get("Status", ""), "title": b.get("Title", ""), "link": b.get("Web Link", "")}
                _store.save("watchlist", wl)
                st.success(f"Watching {b['File']}")
        else:
            addcols[1].caption("Search Albany/Congress first to add those.")
        addcols[2].caption(f"**{len(wl)}** bills on the watchlist.")

    if not wl:
        st.info("Your watchlist is empty. Add a few bills above to start tracking them across levels.")
    else:
        top = st.columns([1, 1, 3])
        if top[0].button("🔄 Refresh & diff", type="primary", key="wl_refresh"):
            changes = []
            for k, w in wl.items():
                new_status = w["last_status"]
                try:
                    if w["level"] == "NYC" and w.get("mid"):
                        det = fetch_detail(w["mid"])
                        his = det.get("histories") or []
                        if his:
                            new_status = (sorted(his, key=lambda h: h.get("MatterHistoryActionDate") or "")[-1]
                                          .get("MatterHistoryActionName") or w["last_status"])
                    elif w["level"] == "NY State":
                        nkey = st.session_state.get("nys_key", "")
                        if nkey.strip():
                            yr = int(year) if str(year).isdigit() else 2025
                            nb = nys_bill_full(w["file"], yr, nkey.strip())
                            if nb:
                                new_status = nb.get("Status", w["last_status"])
                except Exception:
                    pass
                if new_status and new_status != w["last_status"]:
                    changes.append({"Bill": w["file"], "Level": w["level"], "Was": w["last_status"],
                                    "Now": new_status})
                    w["last_status"] = new_status
            st.session_state["wl_changes"] = changes
            _store.save("watchlist", wl)  # persist any status updates from the diff
        if top[1].button("Clear watchlist", key="wl_clear"):
            st.session_state["watchlist"] = {}; _store.save("watchlist", {}); st.rerun()

        ch = st.session_state.get("wl_changes")
        if ch:
            st.markdown("**What moved since last check:**")
            st.dataframe(pd.DataFrame(ch), hide_index=True, use_container_width=True)
        elif ch == []:
            st.caption("No changes detected on the last refresh.")

        st.markdown("**Watchlist**")
        wdf = pd.DataFrame([{"Bill": w["file"], "Level": w["level"], "Status": w["last_status"],
                             "Title": (w["title"] or "")[:80], "Web Link": w.get("link", "")}
                            for w in wl.values()])
        st.dataframe(wdf, use_container_width=True, height=320,
                     column_config={"Web Link": st.column_config.LinkColumn("Open", display_text="Open")})
        st.caption("This watchlist lives for the session. For durable, restart-proof tracking, the scheduled backend "
                   "(handed to Council IT) keeps a lasting history.")


# ============================================================================
# 🪪 DEEP PROFILE — one polished page per member, any level
# ============================================================================
with t_profile:
    st.subheader("🪪 Member deep profile")
    st.caption("A single, polished page for any official — key facts, committees, record, and an AI 'record at a "
               "glance'. Names and facts come from official sources.")
    p_llm = _get_llm(smart=True)
    lvl = st.radio("Level", ["Federal", "NYC Council", "NY State"], horizontal=True, key="prof_level")

    if lvl == "Federal":
        deleg = load_federal_delegation()
        if not deleg:
            st.info("The delegation roster was unreachable just now; it loads automatically in a normal deployment.")
        else:
            names = {f'{d["name"]} — {d["seat"]}': d for d in deleg}
            who = st.selectbox("Member", list(names.keys()), key="prof_fed")
            d = names[who]
            coms = federal_committees().get(d["extra"].get("bioguide", ""), [])
            import profiles as _profiles
            facts = _profiles.federal_facts(d, committees=coms)
            _render_profile(_profiles, p_llm, "Federal", d["name"], facts, d)
    elif lvl == "NYC Council":
        import profiles as _profiles
        members = get_directory()
        who = st.selectbox("Council Member", members, key="prof_nyc") if members else \
            st.text_input("Member last name", key="prof_nyc_txt")
        if who and st.button("Build profile", key="prof_nyc_go", type="primary"):
            with st.spinner("Assembling record…"):
                try:
                    if bundle and member_bills(rows, who):
                        mb = member_bills(rows, who); stats = dossier_stats(mb, who)
                    else:
                        stats = build_member_dossier(who, year)["stats"]
                    coms = sorted({r.get("Committee/Body") for r in (mb if bundle else [])
                                   if r.get("Committee/Body")})[:6] if bundle else []
                    st.session_state["prof_nyc_data"] = (_profiles.council_facts(who, stats, committees=coms), who)
                except Exception as e:
                    st.error(f"{type(e).__name__}: {e}")
        pd_ = st.session_state.get("prof_nyc_data")
        if pd_ and pd_[1] == who:
            _render_profile(_profiles, p_llm, "NYC", who, pd_[0], {})
    else:  # NY State
        import profiles as _profiles
        nkey = st.session_state.get("nys_key", "")
        if not nkey.strip():
            st.info("Add your **NY State** key on the *State & Federal* tab to load Albany members.")
        else:
            ch = st.radio("Chamber", ["State Senate", "State Assembly"], horizontal=True, key="prof_nys_ch")
            yr = int(year) if str(year).isdigit() else 2025
            try:
                mem = nys_members(yr, nkey.strip(), "SENATE" if ch == "State Senate" else "ASSEMBLY")
            except Exception:
                mem = []
            if mem:
                names = {f'{m["name"]} — District {m["district"]}': m for m in
                         sorted(mem, key=lambda x: (x["district"] or 0))}
                who = st.selectbox("Member", list(names.keys()), key="prof_nys")
                m = names[who]
                _render_profile(_profiles, p_llm, "NY State", m["name"], _profiles.state_facts(m), {})
            else:
                st.info("No members returned (or Albany was unreachable).")


# ============================================================================
# 📦 DISTRICT PACKET — one-click printable briefing bundle
# ============================================================================
with t_packet:
    st.subheader("📦 District Packet")
    st.caption("One click, one printable document: a member's profile, their bills, and the upcoming hearings that "
               "touch their committees — the leave-behind for a meeting or a press hit. Print to PDF from the export.")
    import profiles as _profiles
    pk_llm = _get_llm(smart=True)
    lvl = st.radio("Level", ["NYC Council", "Federal"], horizontal=True, key="pk_level")

    incl = st.columns(3)
    want_glance = incl[0].checkbox("AI 'record at a glance'", value=bool(pk_llm.ready), key="pk_glance")
    want_hear = incl[1].checkbox("Upcoming hearings (45 days)", value=True, key="pk_hear")

    if lvl == "NYC Council":
        members = get_directory()
        who = st.selectbox("Council Member", members, key="pk_member") if members else \
            st.text_input("Member last name", key="pk_member_txt")
        if who and st.button("Build packet", type="primary", key="pk_go_nyc"):
            with st.spinner("Assembling the packet…"):
                try:
                    if bundle and member_bills(rows, who):
                        mb = member_bills(rows, who); stats = dossier_stats(mb, who)
                    else:
                        dd = build_member_dossier(who, year); mb = dd["rows"]; stats = dd["stats"]
                    coms = sorted({r.get("Committee/Body") for r in mb if r.get("Committee/Body")})
                    facts = _profiles.council_facts(who, stats, committees=coms[:6])
                    glance = _profiles.glance(pk_llm, "NYC", who, facts) if want_glance else ""
                    hearings = []
                    if want_hear:
                        today = _dt.date.today()
                        allh = fetch_hearings(str(today), str(today + _dt.timedelta(days=45)))
                        comset = {c.lower() for c in coms}
                        hearings = [h for h in allh if (h.get("Committee / Body") or "").lower() in comset] or allh[:10]
                    md = _packet.build_packet_md(who, "NYC City Council", facts=facts, glance=glance,
                                                 bills=mb, hearings=hearings, as_of=str(_dt.date.today()))
                    st.session_state["packet_out"] = (md, who.replace(" ", "_"), mb, facts)
                except Exception as e:
                    st.error(f"{type(e).__name__}: {e}")
    else:  # Federal
        deleg = load_federal_delegation()
        if not deleg:
            st.info("Delegation roster unreachable right now; it loads automatically in a normal deployment.")
        else:
            names = {f'{d["name"]} — {d["seat"]}': d for d in deleg}
            pick = st.selectbox("Member", list(names.keys()), key="pk_fed")
            d = names[pick]
            ckey = st.session_state.get("cong_key", "")
            if st.button("Build packet", type="primary", key="pk_go_fed"):
                with st.spinner("Assembling the packet…"):
                    coms = federal_committees().get(d["extra"].get("bioguide", ""), [])
                    spon = []
                    if ckey.strip():
                        try:
                            spon = congress_member_bills(d["extra"].get("bioguide", ""), "sponsored", ckey.strip())
                        except Exception:
                            spon = []
                    facts = _profiles.federal_facts(d, committees=coms, sponsored=spon)
                    glance = _profiles.glance(pk_llm, "Federal", d["name"], facts) if want_glance else ""
                    md = _packet.build_packet_md(d["name"], "U.S. Congress", facts=facts, glance=glance,
                                                 bills=spon, as_of=str(_dt.date.today()))
                    st.session_state["packet_out"] = (md, d["name"].replace(" ", "_"), spon, facts)

    out = st.session_state.get("packet_out")
    if out:
        md, title, bills_, facts_ = out
        st.markdown(f'<div class="brief">{_brief.md_to_html(md)}</div>', unsafe_allow_html=True)
        st.divider()
        dc = st.columns(3)
        dc[0].download_button("⬇️ Markdown", md, f"packet_{title}.md", "text/markdown",
                              key="pk_md", use_container_width=True)
        html = _brief.print_html(_brief.md_to_html(md), title=f"District packet — {title}")
        dc[1].download_button("🖨️ Print / PDF (HTML)", html, f"packet_{title}.html", "text/html",
                              key="pk_html", use_container_width=True)
        try:
            from openpyxl import Workbook as _WB3
            wb = _WB3(); ws = wb.active; ws.title = "Bills"
            hdrs = ["File", "Type", "Title", "Status", "Prime sponsor", "Web Link"]
            ws.append(hdrs)
            for r_ in _packet.packet_to_rows(title, facts_, bills_):
                ws.append([r_.get(h, "") for h in hdrs])
            wb.save("/tmp/packet.xlsx")
            with open("/tmp/packet.xlsx", "rb") as fh:
                dc[2].download_button("⬇️ Bills (Excel)", fh.read(), f"packet_{title}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="pk_xlsx", use_container_width=True)
        except Exception:
            dc[2].caption("Excel export unavailable.")


# ============================================================================
# 📣 POLITICS & MESSAGING — the member's own communications & influence layer
# (advocacy in the member's voice; grounded, never fabricates stats or quotes)
# ============================================================================
_POLI_NOTE = ("This layer writes in **the member's own voice** — it's advocacy, not the app's neutral analysis. "
              "It never invents statistics or quotes: give it the verified facts and it builds the message around "
              "them, flagging anything to confirm as `[verify]`.")

with t_statement:
    st.subheader("📝 Statement Studio")
    st.caption(_POLI_NOTE)
    s_llm = _get_llm(smart=True)
    if not s_llm.ready:
        st.info("💡 Add your **Anthropic key** in the ⚙️ controls panel to draft statements.")
    with st.expander("✨ Style anchor — the register these drafts mirror"):
        st.markdown(f"> {_msg.EXEMPLAR}")
        st.caption("Firm, values-forward, no invented numbers. Toggle it off below to use a neutral register.")
    issue = st.text_input("Issue / topic", placeholder="e.g. NYPD headcount and the response to sexual-violence data",
                          key="ss_issue")
    stance = st.text_area("What the member wants to convey (their position)", height=80, key="ss_stance",
        placeholder="e.g. We won budget priorities but must restore NYPD headcount; sexual violence can't be "
                    "reframed by statistics — survivors deserve urgency and real action.")
    facts = st.text_area("Verified facts the member can cite (one per line — used for ALL specifics)", height=90,
        key="ss_facts", placeholder="Budget restored X (confirm)\nNYPD headcount down vs. promise (confirm figure)\n"
                                     "Reported cases / category change — cite DCJS/NYPD source")
    _grounded_figures_panel("ss")
    with st.expander("💵 Budget & headcount — authoritative sources to cite"):
        st.caption("Budget and agency-headcount figures are nuanced and easy to get wrong, so this tool doesn't "
                   "auto-generate them. Pull the exact number from an official source below, then paste it into the "
                   "facts box with its citation.")
        st.markdown(
            "- **NYC OMB — Adopted Budget & Financial Plan:** https://www.nyc.gov/site/omb/publications/publications.page\n"
            "- **Council Finance Division (budget reports):** https://council.nyc.gov/budget/\n"
            "- **Independent Budget Office (IBO):** https://www.ibo.nyc.ny.us\n"
            "- **Checkbook NYC (spending & budget data):** https://www.checkbooknyc.com\n"
            "- **Citywide budgeted headcount (NYC Open Data):** https://data.cityofnewyork.us (search “headcount”)")
    c = st.columns(3)
    fmt = c[0].selectbox("Format", list(_msg.FORMATS.keys()), key="ss_fmt")
    tone = c[1].selectbox("Tone", list(_msg.TONES.keys()), index=1, key="ss_tone")
    use_ex = c[2].checkbox("Mirror the style anchor", value=True, key="ss_ex")
    if st.button("Draft it", type="primary", key="ss_go"):
        if not s_llm.ready:
            st.warning("Add your Anthropic key in the ⚙️ controls panel first.")
        elif not issue.strip():
            st.warning("Enter an issue/topic.")
        else:
            with st.spinner("Drafting…"):
                factlist = [x for x in facts.splitlines() if x.strip()]
                st.session_state["ss_out"] = _msg.draft_statement(
                    s_llm, issue.strip(), stance.strip(), factlist, fmt=fmt, tone=tone, use_exemplar=use_ex)
    if st.session_state.get("ss_out"):
        out = st.session_state["ss_out"]
        st.markdown(f'<div class="brief">{_brief.md_to_html(out)}</div>', unsafe_allow_html=True)
        st.code(out, language="markdown")
        st.download_button("⬇️ Download", out, "statement.md", "text/markdown", key="ss_dl")
        st.caption("Draft for staff review. Confirm any `[verify]` figure against DCJS / NYPD / OMB / IBO before release.")

with t_rapid:
    st.subheader("⚡ Rapid Response")
    st.caption(_POLI_NOTE + "  \nPaste what was said (you supply it — the tool never manufactures anyone's quote) "
               "and get a grounded, on-message reply that answers the substance.")
    r_llm = _get_llm(smart=True)
    if not r_llm.ready:
        st.info("💡 Add your **Anthropic key** in the ⚙️ controls panel to use Rapid Response.")
    claim = st.text_area("What was said (the claim to answer)", height=90, key="rr_claim",
        placeholder="e.g. A public official argues a rise in a crime category is a statistical technicality.")
    who = st.text_input("Said by (role — optional)", key="rr_who", placeholder="e.g. the Mayor")
    position = st.text_area("The member's position", height=70, key="rr_pos")
    rfacts = st.text_area("Verified facts to cite (one per line)", height=70, key="rr_facts")
    _grounded_figures_panel("rr")
    rc = st.columns(2)
    rfmt = rc[0].selectbox("Format", list(_msg.FORMATS.keys()), index=1, key="rr_fmt")
    rtone = rc[1].selectbox("Tone", list(_msg.TONES.keys()), index=1, key="rr_tone")
    if st.button("Draft response", type="primary", key="rr_go"):
        if not r_llm.ready:
            st.warning("Add your Anthropic key first.")
        elif not claim.strip():
            st.warning("Paste the statement you're responding to.")
        else:
            with st.spinner("Drafting response…"):
                st.session_state["rr_out"] = _msg.rebuttal(
                    r_llm, claim.strip(), position.strip(), who=who.strip(),
                    facts=[x for x in rfacts.splitlines() if x.strip()], fmt=rfmt, tone=rtone)
    if st.session_state.get("rr_out"):
        out = st.session_state["rr_out"]
        st.markdown(f'<div class="brief">{_brief.md_to_html(out)}</div>', unsafe_allow_html=True)
        st.code(out, language="markdown")
        st.download_button("⬇️ Download", out, "response.md", "text/markdown", key="rr_dl")
        st.caption("Answers the position and the record — not the person. Verify any figure before release.")

with t_influence:
    st.subheader("🧭 Influence Map")
    st.caption("Where a majority comes from on a given issue — across the progressive wing, the moderates, and the "
               "Republican minority (and cross-cutting caucuses), grounded in who actually co-sponsors with whom.")
    i_llm = _get_llm(smart=True)
    with st.expander("🏛️ The Council's blocs (structure — verify current rosters)"):
        for f in _people.COUNCIL_FACTIONS:
            st.markdown(f"- **{f['name']}** ({f['kind']}) — {f['note']}")
    issue_i = st.text_input("Issue", key="im_issue", placeholder="e.g. restoring NYPD headcount in the next budget")
    goal_i = st.text_input("The member's goal", key="im_goal",
                           placeholder="e.g. build a veto-proof majority for a headcount restoration")
    ci = st.columns(2)
    allow_web = ci[0].checkbox("🌐 Allow web search (current caucus rosters, recent positions)", value=True, key="im_web")
    use_coms = ci[1].checkbox("🏛️ Include official committee leadership (loads from Legistar)", value=True, key="im_coms")
    committees_i = []
    if use_coms:
        with st.spinner("Loading committee chairs & members from Legistar…"):
            committees_i = get_committees()
        if committees_i:
            with st.expander(f"🏛️ Committee leadership — {len(committees_i)} committees (official, current)"):
                st.dataframe(pd.DataFrame([{"Committee": c["committee"], "Chair": c["chair"] or "—",
                    "Members": len(c["members"])} for c in committees_i]),
                    hide_index=True, use_container_width=True, height=280)
            st.caption("✓ The memo will weigh committee chairs — the real gatekeepers for moving a bill.")
        else:
            st.caption("Committee roster unreachable right now (loads in a normal deployment).")
    coal_evidence = {}
    if bundle and any(r.get("_sponsor_names") for r in rows):
        try:
            members_c, deg_c, pair_c = coalition_matrix(rows, top_members=20)
            coal_evidence = {"most_active": [{"member": m, "bills": deg_c.get(m, 0)} for m in members_c[:20]],
                             "top_partnerships": sorted(
                                 ([{"a": a, "b": b, "shared": w} for (a, b), w in pair_c.items()]),
                                 key=lambda x: -x["shared"])[:20]}
        except Exception:
            coal_evidence = {}
        st.caption("✓ Using co-sponsorship coalitions from the loaded legislation as evidence.")
    else:
        st.caption("Load legislation **with sponsors** (⚙️ panel) to ground this in real co-sponsorship coalitions; "
                   "it still works from general dynamics without them.")
    if st.button("Build influence memo", type="primary", key="im_go"):
        if not i_llm.ready:
            st.warning("Add your Anthropic key in the ⚙️ controls panel first.")
        elif not issue_i.strip():
            st.warning("Enter an issue.")
        else:
            with st.spinner("Analyzing the landscape…"):
                coms_summary = [{"committee": c["committee"], "chair": c["chair"]}
                                for c in committees_i] if committees_i else []
                swing = {}
                if bundle and any(r.get("_sponsor_names") for r in rows):
                    swing = _analysis.swing_members(rows, issue_i.strip(), committees=committees_i, top=12)
                    st.session_state["im_swing"] = swing
                pers = [{"member": c["member"], "why": "; ".join(c["reasons"])}
                        for c in swing.get("candidates", [])] if swing else []
                st.session_state["im_out"] = _msg.influence_memo(
                    i_llm, issue_i.strip(), goal_i.strip(), _people.factions_reference_text(),
                    coalitions=coal_evidence, committees=coms_summary, persuadables=pers,
                    allow_web=allow_web)
    sw = st.session_state.get("im_swing")
    if sw and sw.get("candidates"):
        st.markdown("**🎯 Computed swing / persuadable members** "
                    f"{'(committees: ' + ', '.join(sw['matched_committees']) + ')' if sw.get('matched_committees') else ''}")
        st.dataframe(pd.DataFrame([{"Member": c["member"], "Score": c["score"],
            "On committee": "✓" if c["on_committee"] else "", "Topic bills": c["topic_bills"],
            "Why": "; ".join(c["reasons"])} for c in sw["candidates"]]),
            hide_index=True, use_container_width=True, height=320)
        st.caption(sw["note"])
    if st.session_state.get("im_out"):
        out = st.session_state["im_out"]
        st.markdown(f'<div class="brief">{_brief.md_to_html(out)}</div>', unsafe_allow_html=True)
        st.download_button("⬇️ Download memo", out, "influence_memo.md", "text/markdown", key="im_dl")
        st.caption("Strategic inference from coalition data + general dynamics — verify roster/vote specifics before "
                   "acting. Members named only where evidence or a source supports it.")


# ============================================================================
# 📖 CM WIKI — a personality-driven page per Council Member (grounded in record)
# ============================================================================
with t_wiki:
    st.subheader("📖 Council Member Wiki")
    st.caption("A living, record-driven page for each Council Member: their legislative persona, the decisions they've "
               "made, who they build with, and an **estimated lean** on any issue. Everything is drawn from their "
               "public sponsorship record — persona describes legislative *style* (not the person), and lean is an "
               "inference from what they've sponsored, **not** a prediction of their vote.")
    import profiles as _profiles
    w_llm = _get_llm(smart=True)
    if not (bundle and any(r.get("_sponsor_names") for r in rows)):
        st.info("Load legislation **with sponsors** (⚙️ panel: *All legislation* + *Include sponsors*) so the Wiki can "
                "read each member's record.")
    else:
        allm = list(_analysis.member_names(rows).keys())
        who = st.selectbox("Council Member", allm, key="wiki_member")
        if who:
            _mem().log("view", "member", who)
            mnc = st.columns([1, 4])
            if mnc[0].button(("★ Following" if _mem().is_following("member", who) else "☆ Follow"),
                             key="wiki_follow"):
                (_mem().unfollow if _mem().is_following("member", who) else _mem().follow)("member", who)
                st.rerun()
            mb = member_bills(rows, who)
            stats = dossier_stats(mb, who)
            _p3 = sum(1 for r in mb if who.split()[-1].lower() in (r.get("Prime Sponsor", "") or "").lower())
            # header
            st.markdown(f'<div class="card pcard"><h4>{_level_badge("NYC")} {who}</h4>'
                        f'<div class="meta">NYC City Council · {stats.get("bills_on",0)} bills on record</div></div>',
                        unsafe_allow_html=True)
            m = st.columns(4)
            m[0].metric("Bills on", stats.get("bills_on", 0))
            m[1].metric("As prime", stats.get("as_prime", _p3))
            m[2].metric("As co-sponsor", stats.get("as_cosponsor", 0))
            m[3].metric("Passed/enacted", (stats.get("by_status") or {}).get("passed", 0))

            st.markdown("### 🎭 Legislative persona")
            if w_llm.ready:
                pk = f"persona::{who}"
                if pk not in st.session_state:
                    with st.spinner("Reading the record…"):
                        st.session_state[pk] = _profiles.persona(w_llm, who, stats)
                if st.session_state.get(pk):
                    st.markdown(f'<div class="brief">{_brief.md_to_html(st.session_state[pk])}</div>',
                                unsafe_allow_html=True)
                st.caption("AI characterization of legislative style from public sponsorship data — not personal "
                           "traits, not an official position.")
            else:
                st.caption("Add your Anthropic key (⚙️ panel) for the AI persona. The facts below work without it.")

            cwa, cwb = st.columns(2)
            if stats.get("by_topic"):
                cwa.markdown("**Policy focus**")
                cwa.bar_chart(pd.Series(dict(list(stats["by_topic"].items())[:8])))
            if stats.get("top_coalition"):
                cwb.markdown("**Builds with**")
                cwb.bar_chart(pd.Series(stats["top_coalition"]))

            st.markdown("### 🗳️ Decisions — prime-sponsored bills")
            pm = [r for r in mb if who.split()[-1].lower() in (r.get("Prime Sponsor", "") or "").lower()]
            if pm:
                st.dataframe(pd.DataFrame([{"File": r["File"], "Type": r["Type"], "Title": (r["Title"] or "")[:80],
                    "Status": r["Status"], "Web Link": r["Web Link"]} for r in pm]),
                    use_container_width=True, height=280,
                    column_config={"Web Link": st.column_config.LinkColumn("Open", display_text="Open")})
            else:
                st.caption("No prime-sponsored bills in the loaded set.")

            st.markdown("### 🎯 Estimate their lean on an issue")
            iss = st.text_input("Issue / policy perspective", key="wiki_issue",
                                placeholder="e.g. tenant protections · police oversight · e-bike safety")
            blend_votes = st.checkbox("Also check how they actually voted on related bills (slower — pulls "
                                      "roll-calls)", key="wiki_blend")
            if iss.strip():
                est = _analysis.estimate_lean(rows, who, iss.strip())
                sig = est["signal"]
                if blend_votes:
                    topic_bills = [r for r in mb if _analysis.topic_match(r, iss.strip())]
                    voted = [r for r in topic_bills
                             if any(w in (r.get("Status", "") or "").lower()
                                    for w in ("enact", "adopt", "approv", "passed"))][:8]
                    vevents = []
                    with st.spinner(f"Pulling roll-calls for {len(voted)} related bill(s)…"):
                        for r in voted:
                            try:
                                vevents += fetch_votes(r["MatterId"])
                            except Exception:
                                pass
                    vc = _analysis.vote_signal(vevents, who)
                    est = _analysis.blend_lean(est, vc)
                lean_label = est.get("blended", est["lean"])
                is_support = any(w in lean_label.lower() for w in ("support", "yes", "leads"))
                badge = "b-green" if is_support else ("b-fed" if "oppos" in lean_label.lower() else "b-muted")
                st.markdown(f'{_badge(lean_label, badge)} &nbsp; confidence: **{est["confidence"]}**',
                            unsafe_allow_html=True)
                st.caption(f"Sponsorship: {sig['on_topic']} related bill(s) — {sig['as_prime']} prime, "
                           f"{sig['as_cosponsor']} co-sponsor. "
                           + (f"e.g. {', '.join(sig['examples'])}. " if sig["examples"] else "")
                           + (f"Actual votes on related bills: {est['vote_counts']['aye']} yes / "
                              f"{est['vote_counts']['nay']} no." if est.get("vote_counts") else ""))
                st.warning("⚠️ " + est["caveat"])


# ============================================================================
# 🎯 ISSUE WAR ROOM — one topic → briefing + figures + swing + memo + statement
# ============================================================================
with t_warroom:
    st.subheader("🎯 Issue War Room")
    st.caption("One topic in, a full kit out: a briefing on what's moving, grounded figures to cite, the swing members "
               "to work, an influence memo, and a draft statement — assembled in one pass and exportable together.")
    wr_llm = _get_llm(smart=True)
    topic = st.text_input("Issue / topic", key="wr_topic",
                          placeholder="e.g. restoring NYPD headcount and responding to sexual-violence data")
    wc = st.columns(2)
    stance = wc[0].text_input("The member's stance (for the draft statement)", key="wr_stance")
    goal = wc[1].text_input("The member's goal (for the influence memo)", key="wr_goal")
    _grounded_figures_panel("wr")
    if not wr_llm.ready:
        st.info("💡 Add your **Anthropic key** (⚙️ panel) for the briefing, memo, and statement. Swing members and "
                "figures work without it.")
    if st.button("🚀 Assemble war room", type="primary", key="wr_go"):
        if not topic.strip():
            st.warning("Enter an issue/topic.")
        else:
            out = {}
            with st.spinner("Assembling…"):
                # 1) swing members (data only)
                if bundle and any(r.get("_sponsor_names") for r in rows):
                    coms = get_committees() if st.session_state.get("im_coms", True) else []
                    out["swing"] = _analysis.swing_members(rows, topic.strip(), committees=coms, top=10)
                # 2) briefing on the topic (from loaded bills)
                if wr_llm.ready and bundle and rows:
                    match = [r for r in rows if _analysis.topic_match(r, topic.strip())]
                    ev = {"topic": topic.strip(), "matching_count": len(match),
                          "sample": [{"File": r["File"], "Title": (r["Title"] or "")[:90], "Status": r["Status"],
                                      "Prime": r.get("Prime Sponsor", "")} for r in match[:25]]}
                    out["briefing"] = _brief.topic_briefing(wr_llm, topic.strip(), ev)
                # 3) influence memo (with committees + swing)
                if wr_llm.ready:
                    coms_sum = [{"committee": c["committee"], "chair": c["chair"]}
                                for c in (out.get("swing") and get_committees() or [])]
                    coal_ev = {}
                    if bundle and any(r.get("_sponsor_names") for r in rows):
                        try:
                            mm, dd, pp = coalition_matrix(rows, top_members=18)
                            coal_ev = {"most_active": [{"member": x, "bills": dd.get(x, 0)} for x in mm[:18]]}
                        except Exception:
                            coal_ev = {}
                    pers = [{"member": c["member"], "why": "; ".join(c["reasons"])}
                            for c in out.get("swing", {}).get("candidates", [])]
                    out["memo"] = _msg.influence_memo(wr_llm, topic.strip(), goal.strip(),
                        _people.factions_reference_text(), coalitions=coal_ev, committees=coms_sum,
                        persuadables=pers, allow_web=True)
                # 4) draft statement
                if wr_llm.ready:
                    figs = st.session_state.get("wr_snap") or []
                    factlist = [f["citation"] for f in figs]
                    out["statement"] = _msg.draft_statement(wr_llm, topic.strip(),
                        stance.strip() or f"The member's priorities on {topic.strip()}", factlist,
                        fmt="Press statement", tone="Firm")
            st.session_state["wr_out"] = out

    wr = st.session_state.get("wr_out")
    if wr:
        if wr.get("swing", {}).get("candidates"):
            st.markdown("### 🎯 Swing members to work")
            st.dataframe(pd.DataFrame([{"Member": c["member"], "Score": c["score"],
                "On committee": "✓" if c["on_committee"] else "", "Why": "; ".join(c["reasons"])}
                for c in wr["swing"]["candidates"]]), hide_index=True, use_container_width=True, height=280)
            st.caption(wr["swing"]["note"])
        if wr.get("briefing"):
            st.markdown("### 📰 Briefing")
            st.markdown(f'<div class="brief">{_brief.md_to_html(wr["briefing"])}</div>', unsafe_allow_html=True)
        if wr.get("memo"):
            st.markdown("### 🧭 Influence memo")
            st.markdown(f'<div class="brief">{_brief.md_to_html(wr["memo"])}</div>', unsafe_allow_html=True)
        if wr.get("statement"):
            st.markdown("### 📝 Draft statement")
            st.markdown(f'<div class="brief">{_brief.md_to_html(wr["statement"])}</div>', unsafe_allow_html=True)
        # combined export
        parts = [f"# Issue War Room — {topic.strip()}", ""]
        if wr.get("briefing"): parts += ["## Briefing", wr["briefing"], ""]
        if wr.get("swing", {}).get("candidates"):
            parts += ["## Swing members to work"]
            parts += [f"- **{c['member']}** (score {c['score']}) — {'; '.join(c['reasons'])}"
                      for c in wr["swing"]["candidates"]]
            parts += [""]
        if wr.get("memo"): parts += ["## Influence memo", wr["memo"], ""]
        if wr.get("statement"): parts += ["## Draft statement", wr["statement"], ""]
        combined = "\n".join(parts)
        st.download_button("⬇️ Download the whole kit (Markdown)", combined,
                           f"war_room_{topic.strip()[:30].replace(' ','_')}.md", "text/markdown", key="wr_dl")
        st.caption("Drafts and estimates are decision-support — verify figures and confirm positions before acting.")


# ============================================================================
# 📊 POLICY GRID — every member × every policy area, at a glance
# ============================================================================
with t_grid:
    st.subheader("📊 Policy Grid — member × policy area")
    st.caption("Each member's whole topic portfolio at once: how many bills they've sponsored in every policy area. "
               "Darker = more bills. The clearest single view of who owns which issue.")
    if not (bundle and any(r.get("_sponsor_names") for r in rows)):
        st.info("Load legislation **with sponsors** (⚙️ panel) so the grid can read every member's record.")
    else:
        topn = st.slider("Members to show (most active first)", 8, 40, 20, key="grid_top")
        members_g, topics_g, mat = _analysis.engagement_matrix(rows, top_members=topn)
        if not members_g or not topics_g:
            st.warning("No topic data in the loaded set yet.")
        else:
            def _short(n):
                return n.split()[-1] if n.split() else n
            idx, seen = [], {}
            for m in members_g:
                seen[_short(m)] = seen.get(_short(m), 0) + 1
            for m in members_g:
                idx.append(m if seen[_short(m)] > 1 else _short(m))
            M = pd.DataFrame([[mat[m][t] for t in topics_g] for m in members_g], index=idx, columns=topics_g)
            mx = max(1, int(M.values.max()))

            def _cell(v):
                if not v:
                    return "background-color:#0a0f1c;color:#26344d"
                t = v / mx
                rr = int(15 + t * (59 - 15)); gg = int(26 + t * (130 - 26)); bb = int(50 + t * (246 - 50))
                return f"background-color:rgb({rr},{gg},{bb});color:#eaf1fb"
            styled = M.style
            styled = (styled.map(_cell) if hasattr(styled, "map") else styled.applymap(_cell))
            styled = styled.format(lambda v: "" if not v else int(v))
            st.dataframe(styled, use_container_width=True, height=min(720, 80 + 26 * len(members_g)))
            st.caption(f"Darkest cell = {mx} bills. Columns are the auto-tagged policy topics; rows are the most "
                       "active sponsors. Click a column header to sort.")
            # who owns each topic
            leaders = []
            for t in topics_g:
                best = max(members_g, key=lambda m: mat[m][t])
                if mat[best][t]:
                    leaders.append({"Policy area": t, "Most active member": best, "Bills": mat[best][t]})
            if leaders:
                st.markdown("**Who owns each issue (most bills sponsored):**")
                st.dataframe(pd.DataFrame(sorted(leaders, key=lambda x: -x["Bills"])),
                             hide_index=True, use_container_width=True, height=300)


# ============================================================================
# 🏛️ CITY HALL — officials, council members (with photos), district profiles
# ============================================================================
@st.cache_data(ttl=86400, show_spinner=False)
def member_photo_cached(name):
    if not _media:
        return None, None
    return _media.wiki_photo(name, "New York City Council")


def _portrait(name, photo_url=None, size=64):
    if _media:
        return _media.portrait_html(name, photo_url, size)
    return ""


with t_officials:
    st.subheader("🏛️ Who's in what office — City of New York")
    st.caption("The citywide and boroughwide elected offices that run New York City, what each one does, and the "
               "official site for the current officeholder. (Names live on the official pages so this never goes stale.)")
    rows_off = _city.citywide_rows()
    st.markdown("**Citywide**")
    cc = st.columns(2)
    cw = [r for r in rows_off if r["scope"] == "Citywide"]
    for i, r in enumerate(cw):
        cc[i % 2].markdown(
            f'<div class="card pcard"><h4>{r["office"]} '
            f'<span class="badge b-nyc">{r["branch"]}</span></h4>'
            f'<div class="meta">{r["remit"]}<br><a href="{r["link"]}">official site ↗</a></div></div>',
            unsafe_allow_html=True)
    st.markdown("**Borough Presidents**")
    bc = st.columns(5)
    for i, r in enumerate([x for x in rows_off if "Borough President" in x["office"]]):
        bc[i].markdown(f'<div class="card pcard"><h4 style="font-size:.92rem">{r["scope"]}</h4>'
                       f'<div class="meta"><a href="{r["link"]}">office ↗</a></div></div>', unsafe_allow_html=True)
    st.markdown("**District Attorneys**")
    dc = st.columns(5)
    for i, r in enumerate([x for x in rows_off if "District Attorney" in x["office"]]):
        dc[i].markdown(f'<div class="card pcard fed"><h4 style="font-size:.92rem">{r["scope"]}</h4>'
                       f'<div class="meta"><a href="{r["link"]}">office ↗</a></div></div>', unsafe_allow_html=True)
    st.info("The **City Council** (51 members) is on the next tab, with photos. Land-use and budget powers run through "
            "the Council + the Mayor + the Borough Presidents (ULURP) — see the Legislation and Briefings tabs.")


with t_council:
    st.subheader("🧑‍🤝‍🧑 The City Council — 51 members")
    members = get_directory()
    if not members:
        st.info("The Council roster loads from Legistar; it was unreachable just now, or no data is loaded. Open the "
                "⚙️ panel and load any City Council data to populate it.")
    else:
        st.caption(f"{len(members)} current Council Members. Portraits are best-effort from Wikipedia; where there's no "
                   "match you'll see a clean initials avatar. Open any member's full record in **People → CM Wiki**.")
        want_photos = st.checkbox("Fetch member photos from Wikipedia (slower first load)", value=False, key="cm_photos")
        cols = st.columns(3)
        for i, nm in enumerate(members):
            photo = None
            if want_photos:
                try:
                    photo, _ = member_photo_cached(nm)
                except Exception:
                    photo = None
            cols[i % 3].markdown(
                f'<div class="card"><div class="memberrow">{_portrait(nm, photo, 56)}'
                f'<div class="info"><h4>{nm}</h4>'
                f'<div class="meta">{_level_badge("NYC")} City Council</div></div></div></div>',
                unsafe_allow_html=True)


with t_distprofile:
    st.subheader("📍 District Profile")
    st.caption("A map of any Council district, who represents it, and a sourced demographic & language snapshot. "
               "Great for understanding who a policy actually reaches.")
    d = st.selectbox("Council district", list(range(1, 52)), key="dp_district")
    links = _city.district_links(d)
    lc = st.columns([2, 1])
    with lc[0]:
        st.components.v1.html(district_map_html(int(d)), height=420)
    with lc[1]:
        st.markdown(f'<div class="card pcard"><h4>District {d}</h4>'
                    f'<div class="meta">NYC City Council</div></div>', unsafe_allow_html=True)
        st.markdown("**Official links**")
        for label, url in links.items():
            st.markdown(f"- [{label} ↗]({url})")
    st.divider()
    st.markdown("### 👥 Who lives here — demographics & languages")
    dp_llm = _get_llm(smart=True)
    if not dp_llm.ready:
        st.info("Add your **Anthropic key** (⚙️ panel) to generate a sourced demographic & language snapshot "
                "(uses web search). The official links above always work.")
    else:
        if st.button("Build demographic profile", type="primary", key="dp_go"):
            with st.spinner("Researching the district (web-sourced)…"):
                st.session_state[f"dp_prof_{d}"] = _city.district_profile(dp_llm, int(d))
        prof = st.session_state.get(f"dp_prof_{d}")
        if prof:
            st.markdown(f'<div class="brief">{_brief.md_to_html(prof)}</div>', unsafe_allow_html=True)
            st.caption("Web-sourced snapshot — figures carry their source and may lag; verify against the linked "
                       "NYC Planning / Census profiles before citing.")


# ============================================================================
# 🧠 ADAPTIVE INTELLIGENCE — personalized focus, smart search, knowledge base
# ============================================================================
# --- Command Center: "Your focus" (learns from what you use) ---
with t_home:
    st.divider()
    st.markdown('<span class="kicker">Your focus — the desk learns what you work on</span>',
                unsafe_allow_html=True)
    _m = _mem()
    prof = _m.interest_profile()
    fol = _m.follows()
    fc = st.columns(3)
    top_members = prof.get("member", [])
    top_topics = prof.get("topic", [])
    with fc[0]:
        st.markdown("**Most-worked members**")
        if top_members:
            for name, w in top_members[:6]:
                star = "★ " if _m.is_following("member", name) else ""
                st.markdown(f"- {star}{name}")
        else:
            st.caption("Open a member in CM Wiki and it shows up here.")
    with fc[1]:
        st.markdown("**Your top topics**")
        if top_topics:
            st.markdown(" ".join(f'<span class="chip">{t}</span>' for t, _w in top_topics[:8]),
                        unsafe_allow_html=True)
        else:
            st.caption("Search or brief on a topic to build this.")
    with fc[2]:
        st.markdown("**Following**")
        watched = fol.get("member", [])
        if watched:
            for w in watched[:8]:
                st.markdown(f"- ★ {w}")
        else:
            st.caption("Use ☆ Follow on a member's CM Wiki page.")
    rec = _m.recent(8)
    if rec:
        st.caption("Recent: " + " · ".join(f"{r['kind']} {r['entity']}" for r in rec[:8]))
    _st = _m.stats()
    st.caption(f"🧠 Memory: {_st['events']} events · {_st['notes']} notes · {_st['follows']} follows "
               + ("(saved to disk)" if _m.ok else "(session only — disk not writable)"))


# --- Legislation list: fast relevance-ranked "smart search" ---
with t_list:
    if bundle and rows:
        with st.expander("🔎 Smart search — natural-language, relevance-ranked (instant, local)"):
            st.caption("Ask in plain words — e.g. *tenants facing eviction*, *e-bike battery fires*, "
                       "*ferry service to the Rockaways*. Ranks every loaded bill by relevance using a local index "
                       "(no API, no wait).")
            sq = st.text_input("Describe what you're looking for", key="smart_q")
            if sq.strip():
                _mem().log("search", "topic", sq.strip()[:40])
                ix = _get_index(rows)
                hits = ix.search(sq.strip(), top=25)
                st.caption(f"{len(hits)} most relevant of {len(rows)} bills")
                if hits:
                    st.dataframe(pd.DataFrame([{"File": r["File"], "Type": r.get("Type", ""),
                        "Title": (r.get("Title") or "")[:90], "Status": r.get("Status", ""),
                        "Relevance": round(s, 3), "Web Link": r.get("Web Link", "")} for r, s in hits]),
                        use_container_width=True, height=420,
                        column_config={"Web Link": st.column_config.LinkColumn("Open", display_text="Open")})


# --- Knowledge & Memory tab ---
with t_memory:
    st.subheader("🧠 Knowledge & Memory")
    st.caption("The adaptive core: the desk remembers what you work on, and any notes you save here are woven back "
               "into future briefings automatically — your knowledge compounds instead of evaporating.")
    _m = _mem()
    s = _m.stats()
    mc = st.columns(4)
    mc[0].metric("Events learned", s["events"]); mc[1].metric("Knowledge notes", s["notes"])
    mc[2].metric("Saved items", s["saved"]); mc[3].metric("Following", s["follows"])
    if not _m.ok:
        st.warning("Storage isn't writable here, so memory is session-only. In a normal deployment it persists to disk.")

    st.markdown("### 📝 Add a knowledge note")
    st.caption("Saved notes on a **member** or **topic** are auto-injected into that entity's briefings.")
    nc = st.columns([1, 1.4, 3, 1])
    net = nc[0].selectbox("Type", ["member", "topic", "bill"], key="note_etype")
    nent = nc[1].text_input("Who/what (e.g. Hanks, housing, Int 0220-2026)", key="note_entity")
    ntext = nc[2].text_input("Note", key="note_text",
                             placeholder="e.g. Confirmed NYPD headcount figure with Finance on 7/12")
    nc[3].write(""); nc[3].write("")
    if nc[3].button("Save note", key="note_save"):
        if nent.strip() and ntext.strip():
            _m.add_note(net, nent.strip(), ntext.strip()); st.rerun()
        else:
            st.warning("Enter both an entity and a note.")

    notes = _m.all_notes()
    if notes:
        st.markdown("### 📚 Your knowledge base")
        for n in notes[:60]:
            rc = st.columns([5, 1])
            rc[0].markdown(f'<span class="badge b-muted">{n["etype"]}</span> **{n["entity"]}** — {n["note"]}  '
                           f'<span style="color:#8894ab;font-size:.75rem">{n["ts"][:10]}</span>',
                           unsafe_allow_html=True)
            if rc[1].button("Delete", key=f"noted_{n['id']}"):
                _m.delete_note(n["id"]); st.rerun()

    prof = _m.interest_profile()
    if prof:
        st.markdown("### 📈 What the desk has learned about your focus")
        pc = st.columns(3)
        for i, (etype, label) in enumerate([("member", "Members"), ("topic", "Topics"), ("bill", "Bills")]):
            items = prof.get(etype, [])
            with pc[i]:
                st.markdown(f"**{label}**")
                if items:
                    st.dataframe(pd.DataFrame([{label[:-1]: e, "Weight": round(w, 1)} for e, w in items]),
                                 hide_index=True, use_container_width=True)
                else:
                    st.caption("—")

    saved = _m.saved_items()
    if saved:
        st.markdown("### 💾 Saved items")
        for it in saved[:20]:
            with st.expander(f"[{it['kind']}] {it['title']} · {it['ts'][:10]}"):
                st.markdown(it["body"])
